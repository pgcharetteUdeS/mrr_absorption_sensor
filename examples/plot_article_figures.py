"""plot_article_figures.py

Plot figures 3b, 4, 6, 7 from:
    P. Girault et al., "Influence of Losses, Device Size, and Mode Confinement on
    Integrated Micro-Ring Resonator Performance for Absorption Spectroscopy Using
    Evanescent Field Sensing," in Journal of Lightwave Technology, 2022
    doi: 10.1109/JLT.2022.3220982.

Data read from Excel files generated by the mrr_absorption_sensor package:
    - "<mantissa>_2DMAPS_VS_GAMMA_and_R.xlsx"
    - "<mantissa>_ALL_RESULTS.xlsx"

"""
__all__ = [
    "plot_article_figures",
    "plot_article_figures_wrapper",
]

import io
import sys
import argparse
from argparse import ArgumentParser, Namespace
from pathlib import Path
from rich import print, traceback

import matplotlib.pyplot as plt
from matplotlib.axes import Axes
import numpy as np
from openpyxl import load_workbook, Workbook

# rich package initializations
traceback.install(show_locals=True)


def _determine_y_plotting_extrema(
    maps_wb: Workbook,
    results_wb: Workbook,
    results_wb_mrr_sheet_col_names: dict,
) -> tuple[float, float]:
    """
    Determine the maximum values of Smax and alpha*L in the Excel Workbooks

    Args:
        maps_wb (Workbook): Workbook read from "<mantissa>_2DMAPS_VS_GAMMA_and_R.xlsx"
        results_wb (Workbook): Workbook read from "<mantissa>_ALL_RESULTS.xlsx"
        results_wb_mrr_sheet_col_names (dict): MRR sheet column names in
                                              "<mantissa>_ALL_RESULTS.xlsx"


    Returns:
        y_max_s (float): maximum Smax
        y_max_αl (float): maximum alpha*L

    """

    # Calculate y axis limit for alpha*l plots
    αl_max: float = max(
        max(cell.value for cell in row)
        for row in maps_wb["alpha x L (dB)"].iter_rows(min_row=2)
    )
    y_max_αl: float = np.ceil(αl_max / 50) * 50

    # Calculate y axis limit for Smax plots
    max_s_max: float = max(
        np.asarray(
            [
                val[0].value
                for val in results_wb["MRR"].iter_rows(
                    min_row=2,
                    min_col=results_wb_mrr_sheet_col_names["maxS_RIU_inv"],
                    max_col=results_wb_mrr_sheet_col_names["maxS_RIU_inv"],
                )
            ]
        )
    )
    y_max_s: float = (
        np.ceil(max_s_max / 50000) * 50000
        if max_s_max < 100000
        else np.ceil(max_s_max / 500000) * 500000
    )

    # Return Smax and alpha*L maxima
    return y_max_s, y_max_αl


def _get_re_rw(results_wb: Workbook, gamma: float) -> tuple[float, float]:
    """
    Get Re and Rw at a given value of gamma

    Args:
        results_wb (Workbook): Workbook read from "<mantissa>_ALL_RESULTS.xlsx"
        gamma (float): gamma value at which Re and Rw are required

    Returns: Re(float), Rw(float)

    """
    gammas: np.ndarray = np.asarray(
        [
            val[0].value
            for val in results_wb["Re and Rw"].iter_rows(min_row=2, max_col=1)
        ]
    )
    index: int = int(np.argmin(np.abs(gammas - gamma)))
    re, rw = np.asarray(
        [
            (val[0].value, val[1].value)
            for val in results_wb["Re and Rw"].iter_rows(
                min_row=2, min_col=3, max_col=4
            )
        ]
    ).T

    return re[index], rw[index]


def _figure_3b(
    results_wb: Workbook,
    out_filename_path: Path,
) -> None:
    """
    Plot alpha_wg(u)
    (figure 3b in the paper)

    Args:
        results_wb (Workbook): Workbook read from "<mantissa>_ALL_RESULTS.xlsx"
        out_filename_path (Path): Output filename Path

    Returns: None

    """

    # fetch u and alpha_wg(u) data
    u, alpha_wg = np.asarray(
        [
            (val[0].value, val[1].value)
            for val in results_wb["alpha_wg_interp"].iter_rows(
                min_row=2, min_col=1, max_col=2
            )
        ]
    ).T

    # Create the figure and plot
    fig, axs = plt.subplots()
    fig.suptitle(f"Figure 3b ('{out_filename_path.stem}*')")
    axs.plot(u, alpha_wg)
    axs.set_ylim(bottom=0)
    axs.set(
        xlabel=(
            "Height (μm)"
            if results_wb["alpha_wg"]["A1"].value == "height_um"
            else "Width (μm)"
        ),
        ylabel=r"$\alpha_{wg}$ (dB/cm)",
    )

    # Save figure to file
    fig.savefig(out_filename_path.parent / f"{out_filename_path.stem}_FIG3b.png")
    print(
        f"Wrote '{str(out_filename_path.parent)}/{out_filename_path.stem}_FIG3b.png'."
    )

    # Explicit None return
    return None


def _figure_4(
    maps_wb: Workbook,
    results_wb: Workbook,
    gamma: float,
    out_filename_path: Path,
) -> None:
    """
    Plot Smax(r), Snr(r), and Se(r) at a given value of gamma
    (figure 4 in the paper)

    Args:
        maps_wb (Workbook): Workbook read from "<mantissa>_2DMAPS_VS_GAMMA_and_R.xlsx"
        results_wb (Workbook): Workbook read from "<mantissa>_ALL_RESULTS.xlsx"
        gamma (float): gamma value for which the plots are required
        out_filename_path (Path): Output filename Path

    Returns: None

    """

    # Fetch r and gamma arrays from their respective sheets in the 2D map Workbook
    r: np.ndarray = np.asarray([c.value for c in maps_wb["R (um)"][1]])
    gammas: np.ndarray = np.asarray(
        [val[0].value for val in maps_wb["gamma (%)"].iter_rows(max_col=1)]
    )

    # Fetch line profiles for Smax, Snr, and Se at gamma from the the 2D map workbook
    index: int = int(np.argmin(np.abs(gammas - gamma)) + 1)
    s: np.ndarray = np.asarray([c.value for c in maps_wb["S (RIU-1)"][index]])
    s_nr: np.ndarray = np.asarray([c.value for c in maps_wb["Snr (RIU-1)"][index]])
    s_e: np.ndarray = np.asarray([c.value for c in maps_wb["Se"][index]])

    # Get corresponding Re and Rw values at gamma from the all_results Workbook
    re, rw = _get_re_rw(
        results_wb=results_wb,
        gamma=gamma,
    )

    # Create the figure
    fig, ax = plt.subplots(constrained_layout=True)
    fig.suptitle(f"Figure 4 ('{out_filename_path.stem}*')")

    # PLot the Smax(r), Snr(r), Se(r) profiles
    ax.semilogx(r, s, "b-", label=r"$S_{MRR}$")
    ax.semilogx(r, s_nr, "r--", label=r"$S_{NR}$")
    index = int(np.argmin(np.abs(r - rw)) + 1)
    ax.semilogx([rw, rw], [0, s_nr[index]], "k")
    ax.text(rw * 0.95, s_nr[index] * 1.05, "RW")
    ax_r: Axes = ax.twinx()
    ax_r.semilogx(r, s_e, "g--", label=r"$S_{e}$")
    index = int(np.argmin(np.abs(r - re)) + 1)
    ax_r.semilogx([re, re], [0, s_e[index]], "k")
    ax_r.text(re * 0.95, s_e[index] * 1.05, "Re")

    # PLot formatting, title, labels, etc.
    y_max: float = (np.ceil((max(s) / 5000)) + 1) * 5000
    ax.set(
        title="Sensitivity components as a function of radius"
        rf" at $\Gamma_{{fluid}}$ = {gamma:.0f}$\%$",
        xlabel="Radius (μm)",
        ylabel=r"$S_{MRR}$ and $S_{NR}$ (RIU$^{-1}$)",
        xlim=(r[0], r[-1]),
        ylim=(0, y_max),
    )
    ax_r.set(ylim=(0, y_max / 1000), ylabel=r"$S_e$")

    # Combine left/right legends
    ax_lines: list = (
        ax.get_legend_handles_labels()[0] + ax_r.get_legend_handles_labels()[0]
    )
    ax_labels: list = (
        ax.get_legend_handles_labels()[1] + ax_r.get_legend_handles_labels()[1]
    )
    ax.legend(ax_lines, ax_labels, loc="upper left")

    # Save figure to file
    fig.savefig(out_filename_path.parent / f"{out_filename_path.stem}_FIG4.png")
    print(f"Wrote '{str(out_filename_path.parent)}/{out_filename_path.stem}_FIG4.png'.")

    # Explicit None return
    return None


def _figure_6_line_profile_plot(
    maps_wb: Workbook,
    results_wb: Workbook,
    gamma: float,
    r: np.ndarray,
    gammas: np.ndarray,
    ax: plt.Axes,
    y_max_s: float,
    y_max_αl: float,
    last: bool,
) -> None:

    """
    PLot a line profile of alpha*L, alpha_prop*L, alpha_bend*L, at gamma in figure 6

    Args:
        maps_wb (Workbook): Workbook read from "<mantissa>_2DMAPS_VS_GAMMA_and_R.xlsx"
        results_wb (Workbook): Workbook read from "<mantissa>_ALL_RESULTS.xlsx"
        gamma (float): gamma value at which the line profiles are required
        r (np.ndarray): array of radii
        gammas (np.ndarray): array of gammas
        ax (Axes): Axes to plot into
        y_max_s (float): Smax maximum value
        y_max_αl (float): alpha*L maximum value
        last (bool): last line profile in the plot (or not)

    Returns: None

    """

    # Calculate the row index in the sheets for the requested gamma value
    index: int = int(np.argmin(np.abs(gammas - gamma)) + 1)

    # Fetch line profiles data from the worksheets
    α_l: np.ndarray = np.asarray([c.value for c in maps_wb["alpha x L (dB)"][index]])
    α_bend_l: np.ndarray = np.asarray(
        [c.value for c in maps_wb["alpha_bend x L (dB)"][index]]
    )
    α_prop_l: np.ndarray = np.asarray(
        [c.value for c in maps_wb["alpha_prop x L (dB)"][index]]
    )
    s: np.ndarray = np.asarray([c.value for c in maps_wb["S (RIU-1)"][index]])

    # Get corresponding Re and Rw values from the all_results workbook
    re, rw = _get_re_rw(
        results_wb=results_wb,
        gamma=gamma,
    )

    # Plot the line profiles
    ax.semilogx(r, α_l, "k", label="αL")
    ax.semilogx(r, α_bend_l, "g--", label="αbendL")
    ax.semilogx(r, α_prop_l, "r--", label="αpropL")
    ax.semilogx([rw, rw], [0, 50], "k")
    ax.text(rw * 1.05, y_max_αl * 1.025, "Rw")
    ax.semilogx([re, re], [0, 50], "k")
    ax.text(re * 0.85, y_max_αl * 1.025, "Re")
    ax_r: Axes = ax.twinx()
    ax_r.semilogx(r, s, "b", label=r"$S_{MRR}$")

    # PLot formatting, title, labels, etc.
    ax.set(
        title=rf"$\Gamma_{{fluid}}$ = {gamma:.0f} $\%$",
        ylabel="Roundtrip losses (dB)",
        xlim=(r[0], r[-1]),
        ylim=(0, y_max_αl),
    )
    if last:
        ax.set_xlabel("Radius (μm)")
    else:
        ax.axes.get_xaxis().set_ticklabels([])
    ax_r.set(ylabel=r"$S_{MRR}$ (RIU$^{-1}$)", ylim=(0, y_max_s))

    # Combine left/right legends
    ax_lines: list = (
        ax.get_legend_handles_labels()[0] + ax_r.get_legend_handles_labels()[0]
    )
    ax_labels: list = (
        ax.get_legend_handles_labels()[1] + ax_r.get_legend_handles_labels()[1]
    )
    ax.legend(ax_lines, ax_labels, loc="upper left")

    # Explicit None return
    return None


def _figure_6(
    maps_wb: Workbook,
    results_wb: Workbook,
    line_profile_gammas: np.ndarray,
    y_max_s: float,
    y_max_αl: float,
    out_filename_path: Path,
) -> None:
    """
    Plot alpha*L, alpha_prop*L, alpha_bend*L, at a list of gamma values
    (figure 6 in the paper)

    Args:
        maps_wb (Workbook): Workbook read from "<mantissa>_2DMAPS_VS_GAMMA_and_R.xlsx"
        results_wb (Workbook): Workbook read from "<mantissa>_ALL_RESULTS.xlsx"
        line_profile_gammas (np.ndarray): gammas at which line profiles are requested
        y_max_s (float): Smax maximum value
        y_max_αl (float): alpha*L maximum value
        out_filename_path (Path): Output filename Path

    Returns: None

    """

    # Fetch r and gamma arrays from their respective sheets in the Workbook
    r: np.ndarray = np.asarray([c.value for c in maps_wb["R (um)"][1]])
    gammas: np.ndarray = np.asarray(
        [val[0].value for val in maps_wb["gamma (%)"].iter_rows(max_col=1)]
    )

    # Create figure with required number of subplots for the requested line profiles
    fig, axs = plt.subplots(
        nrows=len(line_profile_gammas), figsize=(9, 10), constrained_layout=True
    )
    fig.suptitle(f"Figure 6 ('{out_filename_path.stem}*')")

    # Loop to generate the subplots of the line profiles in "line_profile_gammas"
    for i, gamma in enumerate(line_profile_gammas):
        _figure_6_line_profile_plot(
            maps_wb=maps_wb,
            results_wb=results_wb,
            gamma=gamma,
            r=r,
            gammas=gammas,
            ax=axs[i],
            y_max_s=y_max_s,
            y_max_αl=y_max_αl,
            last=i >= len(line_profile_gammas) - 1,
        )

    # Save figure to file
    fig.savefig(out_filename_path.parent / f"{out_filename_path.stem}_FIG6.png")
    print(f"Wrote '{str(out_filename_path.parent)}/{out_filename_path.stem}_FIG6.png'.")

    # Explicit None return
    return None


def _figure_7(
    maps_wb: Workbook,
    results_wb: Workbook,
    results_wb_mrr_sheet_col_names: dict,
    y_max_s: float,
    out_filename_path: Path,
) -> None:
    """
    Plot Smax(r), S(r), Smax(alpha_inv), r(h)
    (figure 7 in the paper)

    Args:
        maps_wb (Workbook): Workbook read from "<mantissa>_2DMAPS_VS_GAMMA_and_R.xlsx"
        results_wb (Workbook): Workbook read from "<mantissa>_ALL_RESULTS.xlsx"
        results_wb_mrr_sheet_col_names (dict): MRR sheet column names
        y_max_s (float): Smax maximum value
        out_filename_path (Path): Output filename Path

    Returns: None

    """

    # Create array of discrete gamma values
    gammas: np.ndarray = np.asarray(
        [val[0].value for val in maps_wb["gamma (%)"].iter_rows(max_col=1)]
    )
    line_profile_gammas: np.ndarray = np.arange(
        np.ceil(gammas[-1] / 10) * 10, np.floor(gammas[0] / 10) * 10 + 10, 10
    )
    if int(gammas[-1]) != line_profile_gammas[0]:
        line_profile_gammas = np.insert(line_profile_gammas, 0, int(gammas[-1]))
    if int(gammas[0]) != line_profile_gammas[-1]:
        line_profile_gammas = np.append(line_profile_gammas, int(gammas[0]))

    # Fetch required arrays from the maps workbook
    maps_r: np.ndarray = np.asarray([c.value for c in maps_wb["R (um)"][1]])

    # Create the figure
    fig, axs = plt.subplots(3, figsize=(9, 12))
    fig.tight_layout(pad=10, w_pad=10, h_pad=10)
    fig.suptitle(f"Figure 7 ('{out_filename_path.stem}*')")

    #
    # 7a
    #

    # Plot Smax(r)
    r: np.ndarray = np.asarray(
        [
            val[0].value
            for val in results_wb["MRR"].iter_rows(
                min_row=2, max_col=results_wb_mrr_sheet_col_names["R_um"]
            )
        ]
    )
    s_max: np.ndarray = np.asarray(
        [
            val[0].value
            for val in results_wb["MRR"].iter_rows(
                min_row=2,
                min_col=results_wb_mrr_sheet_col_names["maxS_RIU_inv"],
                max_col=results_wb_mrr_sheet_col_names["maxS_RIU_inv"],
            )
        ]
    )
    axs[0].loglog(
        r,
        s_max,
        color="blue",
        linewidth=5,
        alpha=0.25,
        label=r"max{S$_{MRR}$}",
    )

    # Loop to plot S(r) at discrete gammas
    for g in line_profile_gammas:
        index: int = int(np.argmin(np.abs(gammas - g)) + 1)
        s: np.ndarray = np.asarray([c.value for c in maps_wb["S (RIU-1)"][index]])
        axs[0].loglog(maps_r, s, label=rf"$\Gamma$ = {g:.0f}%")

    # PLot formatting, title, labels, etc.
    axs[0].set(
        xlabel="Radius (μm)",
        ylabel=r"S$_{MRR}$",
        xlim=(r[0], r[-1]),
        ylim=(10, y_max_s),
    )
    axs[0].legend(loc="lower right", ncol=2)

    #
    # 7b
    #
    for gamma in line_profile_gammas:
        index = int(np.argmin(np.abs(gammas - gamma)) + 1)
        α_inv: np.ndarray = np.asarray(
            [1 / c.value for c in maps_wb["alpha (um-1)"][index]]
        )
        s = np.asarray([c.value for c in maps_wb["S (RIU-1)"][index]])
        axs[1].plot(
            α_inv,
            s,
            label="".join(
                [
                    r"$\Gamma$ = ",
                    f"{gamma:.0f}%",
                    f", R = {maps_r[np.where(s > np.max(s)*0.995)][0]:.0f} μm ",
                    "@ max(Smrr)",
                ]
            ),
        )
    axs[1].set(xlabel=r"1/$\alpha$ ($\mu$m)", ylabel=r"S$_{MRR}$")
    axs[1].legend(loc="upper left", ncol=2)

    #
    # 7c
    #

    # Plot Smax(r)
    axs[2].loglog(r, s_max, color="blue", label=r"max{S$_{MRR}$}")
    axs[2].set(
        xlabel="Radius (μm)",
        ylabel="max(S$_{MRR}$)",
        xlim=(r[0], r[-1]),
        ylim=(10, y_max_s),
    )

    # h(R) @ Smax(r)
    h: np.ndarray = np.asarray(
        [
            val[0].value
            for val in results_wb["MRR"].iter_rows(
                min_row=2,
                min_col=results_wb_mrr_sheet_col_names["h_um"],
                max_col=results_wb_mrr_sheet_col_names["h_um"],
            )
        ]
    )
    y_min_h: float = 0.1
    y_max_h: float = 0.9
    ax_r: Axes = axs[2].twinx()
    ax_r.semilogx(r, h, "black", label="h")
    ax_r.set(
        ylabel=r"h (μm) @ max($S_{MRR}$)",
        xlim=(r[0], r[-1]),
        ylim=(y_min_h, y_max_h),
    )

    # Legend
    ax_lines: list = (
        axs[2].get_legend_handles_labels()[0] + ax_r.get_legend_handles_labels()[0]
    )
    ax_labels: list = (
        axs[2].get_legend_handles_labels()[1] + ax_r.get_legend_handles_labels()[1]
    )
    axs[2].legend(ax_lines, ax_labels, loc="upper left")

    # Save figure to file
    fig.savefig(out_filename_path.parent / f"{out_filename_path.stem}_FIG7.png")
    print(f"Wrote '{str(out_filename_path.parent)}/{out_filename_path.stem}_FIG7.png'.")

    # Explicit None return
    return None


def plot_article_figures(
    results_file_name: str, maps_file_name: str, block: bool = False
) -> None:
    """
    Plot the article figures

    Args:
        results_file_name (str): "<mantissa>_MRR_2DMAPS_VS_GAMMA_and_R.xlsx" filename
        maps_file_name (str): "<mantissa>_ALL_RESULTS.xlsx" filename
        block (bool): figure show blocking/non-blocking

    Returns: None

    """

    # matplotlib initializations
    plt.rcParams.update(
        {
            "axes.grid": True,
            "axes.grid.which": "both",
        },
    )

    # Load Excel workbooks (into memory, to reduce risk of conflict)
    with open(results_file_name, "rb") as f:
        in_mem_results_file = io.BytesIO(f.read())
    with open(maps_file_name, "rb") as f:
        in_mem_maps_file = io.BytesIO(f.read())
    results_wb: Workbook = load_workbook(in_mem_results_file, read_only=True)
    maps_wb: Workbook = load_workbook(in_mem_maps_file, read_only=True)
    out_filename_path: Path = Path(results_file_name)

    # Create dictionary of column names for the MRR worksheet in the Excel results file
    results_wb_mrr_sheet_col_names: dict = {
        col.value: index for index, col in enumerate(results_wb["MRR"][1], start=1)
    }

    # User-defined discrete gamma value array for figure 6
    line_profile_gammas: np.ndarray = np.asarray([20, 45, 65, 75])

    # Generate figures
    y_max_s, y_max_αl = _determine_y_plotting_extrema(
        maps_wb=maps_wb,
        results_wb=results_wb,
        results_wb_mrr_sheet_col_names=results_wb_mrr_sheet_col_names,
    )
    _figure_3b(results_wb=results_wb, out_filename_path=out_filename_path)
    _figure_4(
        maps_wb=maps_wb,
        results_wb=results_wb,
        gamma=30,
        out_filename_path=out_filename_path,
    )
    _figure_6(
        maps_wb=maps_wb,
        results_wb=results_wb,
        line_profile_gammas=line_profile_gammas,
        y_max_s=y_max_s,
        y_max_αl=y_max_αl,
        out_filename_path=out_filename_path,
    )
    _figure_7(
        maps_wb=maps_wb,
        results_wb=results_wb,
        results_wb_mrr_sheet_col_names=results_wb_mrr_sheet_col_names,
        y_max_s=y_max_s,
        out_filename_path=out_filename_path,
    )
    plt.show(block=block)

    # Explicit None return
    return None


def plot_article_figures_wrapper() -> None:
    """
    Wrapper function for calling plot_article_figures() with command line arguments

    Returns: None

    """

    # Parser for command line parameter input (ex: running from .bat file)
    parser: ArgumentParser = argparse.ArgumentParser(description="Plot article figures")
    parser.add_argument("--results_file", type=str)
    parser.add_argument("--maps_file", type=str)
    parser.add_argument("--no_pause", action="store_true")
    args: Namespace = parser.parse_args()

    # Plot article figures
    plot_article_figures(
        results_file_name=args.results_file, maps_file_name=args.maps_file
    )

    # If running from the command line, either pause for user input to keep figures
    # visible. If running in PyCharm debugger, set breakpoint here.
    if sys.gettrace() is not None:
        print("Breakpoint here to keep figures visible in IDE!")
    elif not args.no_pause:
        input("Script paused to display figures, press any key to exit")
    else:
        print("Done!")

    # Explicit None return
    return None


# Run the script...
if __name__ == "__main__":
    plot_article_figures_wrapper()
