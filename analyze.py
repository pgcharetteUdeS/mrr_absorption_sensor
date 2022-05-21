"""

Analyze the 3 sensor types

Exposed methods:
   - analyze()

"""


import sys
from pathlib import Path
from typing import Callable

import colorama as colorama
import matplotlib.pyplot as plt
import numpy as np
from colorama import Fore, Style
from openpyxl.workbook import Workbook

from .constants import constants, __version__
from .fileio import load_toml_file
from .linear import Linear
from .models import Models
from .mrr import Mrr
from .spiral import Spiral


def _validate_excel_output_file(filename_path: Path) -> str:
    """
    Define output Excel filename string from a Path object. Test to see if the file is
    already open, if so return an exception.

    This function is useful to run before any serious works starts because if the script
    tries to open a file that's already open, say in Excel, this causes the script to
    halt with an exception and the work done up to that point is lost.

    Args:
        filename_path (Path): filename path object to test for openness, conversion

    Returns:
        str: filename path string

    """
    excel_output_filename: str = str(
        filename_path.parent / f"{filename_path.stem}_ALL_RESULTS.xlsx"
    )

    try:
        with open(excel_output_filename, "w"):
            pass
    except IOError:
        print(
            f"{Fore.YELLOW}Could not open '{excel_output_filename}', close it "
            + f"if it's already open!{Style.RESET_ALL}"
        )
        sys.exit()

    return excel_output_filename


def _write_excel_results_file(
    excel_output_fname: str,
    models: Models,
    mrr: Mrr,
    linear: Linear,
    spiral: Spiral,
    parameters: dict,
    logger: Callable = print,
):
    """
    Write the analysis results to the output Excel file from a dictionary of
    key:value pairs, where the keys are the Excel file column header text strings
    and the values are the corresponding column data arrays

    Args:
        excel_output_fname (str):
        models (MOdels):
        mrr (Mrr):
        linear (Linear):
        spiral (Spiral):
        parameters (dict):
        logger (Callable):

    Returns: None

    """
    # Create Excel workbook
    wb = Workbook()

    # Save the MMR data to a sheet
    mrr_data_dict = {
        "R_um": models.r,
        "neff": mrr.n_eff,
        "maxS_RIU_inv": mrr.s,
        "Se": mrr.s_e,
        "Snr_RIU_inv": mrr.s_nr,
        "alpha_bend_dB_per_cm": mrr.α_bend * constants.PER_UM_TO_DB_PER_CM,
        "alpha_wg_dB_per_cm": mrr.α_wg * constants.PER_UM_TO_DB_PER_CM,
        "a2": mrr.wg_a2,
        "tau": mrr.tau,
        "T_max": mrr.t_max,
        "T_min": mrr.t_min,
        "ER_dB": mrr.er,
        "contrast": mrr.contrast,
        f"{models.core_u_name}_um": mrr.u,
        "gamma_percent": mrr.gamma,
        "Finesse": mrr.finesse,
        "Q": mrr.q,
        "FWHM_um": mrr.fwhm,
        "FSR_um": mrr.fsr,
    }
    mrr_data: np.ndarray = np.asarray(list(mrr_data_dict.values())).T
    mrr_sheet = wb["Sheet"]
    mrr_sheet.title = "MRR"
    mrr_sheet.append(list(mrr_data_dict.keys()))
    for row in mrr_data:
        mrr_sheet.append(row.tolist())

    # Save the Re(gamma) & Rw(gamma) arrays to a sheet
    re_rw_sheet = wb.create_sheet("Re and Rw")
    re_rw_sheet.append(
        [
            "gamma_percent",
            f"{models.core_u_name}_um",
            "Re_um",
            "Rw_um",
            "A_um_inv",
            "B_um_inv",
        ]
    )
    for line in zip(
        mrr.gamma_resampled * 100,
        mrr.u_resampled,
        mrr.r_e,
        mrr.r_w,
        mrr.α_bend_a,
        mrr.α_bend_b,
    ):
        re_rw_sheet.append(line)

    # Save the linear waveguide data to a sheet
    linear_data_dict = {
        "R_um": models.r,
        "maxS_RIU_inv": linear.s,
        f"{models.core_u_name}_um": linear.u,
        "gamma_percent": linear.gamma,
        "L_um": 2 * models.r,
        "a2": linear.wg_a2,
    }
    linear_data: np.ndarray = np.asarray(list(linear_data_dict.values())).T
    linear_sheet = wb.create_sheet("Linear")
    linear_sheet.append(list(linear_data_dict.keys()))
    for row in linear_data:
        linear_sheet.append(row.tolist())

    # If required, save the spiral data to a sheet
    if not parameters["no_spiral"]:
        spiral_data_dict = {
            "R_um": models.r,
            "maxS_RIU_inv": spiral.s,
            f"{models.core_u_name}_um": spiral.u,
            "gamma_percent": spiral.gamma,
            "n_revs": spiral.n_turns * 2,
            "Rmin_um": spiral.outer_spiral_r_min,
            "L_um": spiral.l,
            "a2": spiral.wg_a2,
        }
        spiral_data: np.ndarray = np.asarray(list(spiral_data_dict.values())).T
        spiral_sheet = wb.create_sheet("Spiral")
        spiral_sheet.append(list(spiral_data_dict.keys()))
        for row in spiral_data:
            spiral_sheet.append(row.tolist())

    # Save the Excel file to disk
    wb.save(filename=excel_output_fname)
    logger(f"Wrote '{excel_output_fname}'.")


def analyze(
    toml_input_file: str,
    logger: Callable = print,
) -> tuple[Models, Mrr | None, Linear | None, Spiral | None]:
    """
    Calculate the maximum achievable sensitivities over a range of radii for micro-ring
    resonator, spiral, and linear waveguide absorption sensors.

    The waveguides core geometry has a fixed dimension "v", the other "u" is allowed
    to vary over a specified range to find the maximum sensitivity at each radius.

    Args:
        toml_input_file (str):
        logger (Callable):

    Returns:
        Models: Models class instance
        Mrr Mrr class instance
        Linear: Linear class instance
        Spiral: Spiral class instance

    """
    # Initialize the colorama package to print colored text to the Windows console
    colorama.init()

    # Show the package version number
    logger(f"{Fore.YELLOW}mrr_absorption_sensor package {__version__}{Style.RESET_ALL}")

    # matplotlib initializations
    plt.rcParams.update(
        {
            "axes.grid": True,
            "axes.grid.which": "both",
        },
    )

    # Load the problem parameters from the input .toml file
    toml_input_file_path: Path = Path(toml_input_file)
    (parameters, modes_data, bending_loss_data,) = load_toml_file(
        filename=toml_input_file_path,
        logger=logger,
    )

    # Build the filename Path object for the output files
    output_filenames_path: Path = (
        toml_input_file_path.parent
        / parameters["output_sub_dir"]
        / toml_input_file_path.name
    )

    # Instantiate the Models class to build/fit the interpolation models for
    # gamma_fluid(u), neff(u), alpha_wg(u), and alpha_bend(u, r)
    models: Models = Models(
        parameters=parameters,
        modes_data=modes_data,
        bending_loss_data=bending_loss_data,
        filename_path=output_filenames_path,
        logger=logger,
    )

    # Check that the array of radii to be analyzed is not empty
    if np.size(models.r) == 0:
        raise ValueError(f"{Fore.YELLOW}No radii to analyze!{Style.RESET_ALL}")

    # If only model fitting was required, return
    if parameters["models_only"]:
        plt.show()
        return models, None, None, None

    # Define output Excel filename: if file is already open, halt with an exception
    # (better to halt here with an exception than AFTER having done the analysis...)
    if parameters["write_excel_files"]:
        excel_output_fname = _validate_excel_output_file(output_filenames_path)
    else:
        excel_output_fname = ""

    # Instantiate sensor classes
    mrr = Mrr(models=models, logger=logger)
    linear = Linear(models=models, logger=logger)
    spiral = Spiral(models=models, logger=logger)

    # Analyze sensors
    mrr.analyze()
    linear.analyze()
    if not parameters["no_spiral"]:
        spiral.analyze()

    # Plot results
    models.calculate_plotting_extrema(max_s=mrr.max_s)
    mrr.plot_optimization_results()
    linear.plot_optimization_results()
    if not parameters["no_spiral"]:
        spiral.plot_optimization_results()
    mrr.plot_combined_linear_mrr_spiral_optimization_results(
        linear=linear, spiral=spiral
    )

    # Write the analysis results to the output Excel file
    if parameters["write_excel_files"]:
        _write_excel_results_file(
            excel_output_fname=excel_output_fname,
            models=models,
            mrr=mrr,
            linear=linear,
            spiral=spiral,
            parameters=parameters,
            logger=logger,
        ),

    # Return the instantiated models class
    plt.show()
    return models, mrr, linear, spiral
