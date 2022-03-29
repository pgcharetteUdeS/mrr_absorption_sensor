"""

File I/O utilities

Exposed methods:
   - load_toml_file()
   - validate_excel_output_file()
   - write_excel_output_file()
   - write_image_data_to_Excel()

All lengths are in units of um

"""


# Standard library
from colorama import Fore, Style
import numpy as np
from openpyxl.workbook import Workbook
from pathlib import Path
import sys
import toml
from typing import Callable

# Package modules
from .models import Models
from .mrr import Mrr
from .linear import Linear
from .spiral import Spiral
from .version import __version__


def _check_mode_solver_data(
    modes_data: dict,
    bending_loss_data: dict,
    filename: Path,
    logger=print,
) -> bool:
    """
    For each u entry in the dictionary, check that the mode solver data are ordered,
    positive, and without duplicates.
    """

    # Check that the mode solver data dictionary is not empty
    if not bending_loss_data:
        logger(f"No bending loss data loaded from '{filename}'!")
        return False

    # Check that u values are in ascending order and positive
    u_bending_loss: np.ndarray = np.asarray(list(bending_loss_data.keys()))
    if not np.all(u_bending_loss[:-1] < u_bending_loss[1:]) or u_bending_loss[0] <= 0:
        logger(f"Bending loss data in '{filename}' are not in ascending h order!")
        return False

    # Check radii and alpha_bend arrays are ordered, positive, and without duplicates
    for u, value in bending_loss_data.items():
        Rs: np.ndarray = np.asarray(value["R"])
        ABs: np.ndarray = np.asarray(value["alpha_bend"])
        if len(Rs) != len(ABs) or len(Rs) < 3:
            logger(f"Invalid R/alpha_bend arrays for u = {u:.3f} in '{filename}'!")
            return False
        if len(np.unique(Rs)) != len(Rs) or not np.all(Rs[:-1] < Rs[1:]) or Rs[0] <= 0:
            logger(f"Invalid R array for u {u:.3f} in '{filename}'!!")
            return False
        if (
            len(np.unique(ABs)) != len(ABs)
            or not np.all(ABs[:-1] > ABs[1:])
            or ABs[-1] <= 0
        ):
            logger(f"Invalid alpha_bend array for u = {u:.3f} in '{filename}'!")
            return False

    # Check that neff and gamma values are reasonable
    for mode in modes_data.values():
        if mode["neff"] < 1 or mode["gamma"] > 1:
            logger(f"Invalid 'modes' data in '{filename}'!")
            return False

    # Return success
    return True


def load_toml_file(
    filename: Path, logger: Callable = print
) -> tuple[dict, dict, dict,]:
    """

    Load problem data from .toml file and parse it into internal dictionaries.

    :param filename: .toml input file containing the problem data
    :param logger: console logger (optional)
    :return: parameters, modes_data, bending_loss_data
             --
             parameters (dict): problem parameters,
             bending_loss_data (dict): R(u) and alpha_bend(u) mode solver data

    """
    # Load dictionary from the .toml file
    toml_data = toml.load(str(filename))

    # Parse fields from .toml file into the parameters{} dictionary
    parameters: dict = {
        # Waveguide physical parameters
        "alpha_wg": toml_data.get("alpha_wg", 1.0),
        "core_height": toml_data.get("core_height"),
        "core_width": toml_data.get("core_width"),
        "lambda_res": toml_data.get("lambda_res", 0.633),
        "ni_op": toml_data.get("ni_op", 1.0e-6),
        "pol": toml_data.get("pol", "TE"),
        # Spiral physical parameters
        "spiral_spacing": toml_data.get("spiral_spacing", 5.0),
        "spiral_turns_min": toml_data.get("spiral_turns_min", 0.5),
        "spiral_turns_max": toml_data.get("spiral_turns_max", 25.0),
        # Model fitting parameters
        "gamma_order": toml_data.get("gamma_order", 4),
        "neff_order": toml_data.get("neff_order", 3),
        # Analysis parameters
        "alpha_bend_threshold": toml_data.get("alpha_bend_threshold", 0.01),
        "min_delta_ni": toml_data.get("min_delta_ni", 1.0e-6),
        "Rmin": toml_data.get("Rmin", 25.0),
        "Rmax": toml_data.get("Rmax", 10000.0),
        "R_samples_per_decade": toml_data.get("R_samples_per_decade", 100),
        "T_SNR": toml_data.get("T_SNR", 20.0),
        # Graphing and file I/O and parameters
        "map2D_colormap": toml_data.get("map2D_colormap", "viridis"),
        "map2D_overlay_color_dark": toml_data.get("map2D_overlay_color_dark", "white"),
        "map2D_overlay_color_light": toml_data.get(
            "map2D_overlay_color_light", "black"
        ),
        "map_line_profiles": toml_data.get("map_line_profiles", []),
        "output_sub_dir": toml_data.get("output_sub_dir", ""),
        "write_excel_files": toml_data.get("write_excel_files", True),
        "write_spiral_sequence_to_file": toml_data.get(
            "write_spiral_sequence_to_file", True
        ),
        # Debugging and other flags
        "disable_u_search_lower_bound": toml_data.get(
            "disable_u_search_lower_bound", False
        ),
        "disable_R_domain_check": toml_data.get("disable_R_domain_check", False),
        "models_only": toml_data.get("models_only", False),
        "no_spiral": toml_data.get("no_spiral", False),
    }

    # Check if .toml file contains unsupported keys, if so exit
    valid_keys: list = list(parameters.keys()) + ["h", "w"]
    invalid_keys: list = [key for key in toml_data.keys() if key not in valid_keys]
    if invalid_keys:
        valid_keys.sort(key=lambda x: x.lower())
        logger(
            f"{Fore.YELLOW}File '{filename}' contains unsupported keys: "
            + f"{invalid_keys} - Valid keys are : {valid_keys}{Style.RESET_ALL}"
        )
        sys.exit()

    # Determine if this is a "fixed core height" or "fixed core width" analysis
    if (
        toml_data.get("core_width") is not None
        and toml_data.get("core_height") is not None
    ) or (toml_data.get("core_width") is None and toml_data.get("core_height") is None):
        logger(
            f"{Fore.YELLOW}EITHER 'core_width' or 'core_height' fields "
            + f"nust be specified!{Style.RESET_ALL}"
        )
        sys.exit()
    if toml_data.get("core_width") is not None:
        logger("Fixed waveguide core width analysis.")
        parameters["core_u_name"] = "h"
        parameters["core_v_name"] = "w"
        parameters["core_v_value"] = toml_data.get("core_width")
        if toml_data.get("h") is None:
            logger(f"{Fore.YELLOW}No 'h' fields specified!{Style.RESET_ALL}")
            sys.exit()
    else:
        logger("Fixed waveguide core height analysis.")
        parameters["core_u_name"] = "w"
        parameters["core_v_name"] = "h"
        parameters["core_v_value"] = toml_data.get("core_height")
        if toml_data.get("w") is None:
            logger(f"{Fore.YELLOW}No 'w' fields specified!{Style.RESET_ALL}")
            sys.exit()

    # Check selected keys for valid content
    if parameters["pol"] not in ["TE", "TM"]:
        logger(
            f"{Fore.YELLOW}Invalid 'pol' field value '{parameters['pol']}'"
            + f" in '{filename}'!{Style.RESET_ALL}"
        )
        sys.exit()

    # Copy the neff(u) and gamma(u) mode solver data to the "modes_data{}" dictionary,
    # and the alpha_bend(R, u) mode solver data to the "bending_loss_data{}" dictionary
    modes_data: dict = {}
    bending_loss_data: dict = {}
    for k, value in toml_data[parameters["core_u_name"]].items():
        u_key_um = float(k) / 1000
        modes_data[u_key_um] = {
            "u": u_key_um,
            "neff": value["neff"],
            "gamma": value["gamma"],
        }
        bending_loss_data[u_key_um] = {
            "R": value["R"],
            "alpha_bend": value["alpha_bend"],
        }

    # Check the validity of the mode solver data, exit if problem found
    if not _check_mode_solver_data(
        modes_data=modes_data,
        bending_loss_data=bending_loss_data,
        filename=filename,
        logger=logger,
    ):
        logger(f"{Fore.YELLOW}Mode solver data validation fail!{Style.RESET_ALL}")
        sys.exit()

    # Console message
    logger(f"Loaded information from '{filename}'.")

    # Write out the parameters dictionary to a text file
    filename_txt: Path = (
        filename.parent
        / parameters["output_sub_dir"]
        / f"{filename.stem}_parameters.toml"
    )
    with open(filename_txt, "w") as f:
        f.write(f"# mrr_absorption_sensor package {__version__}\n")
        toml.dump(parameters, f)
    logger(f"Wrote input parameters to '{filename_txt}'.")

    return (
        parameters,
        modes_data,
        bending_loss_data,
    )


def validate_excel_output_file(filename_path: Path) -> str:
    """
    Define output Excel filename string from a Path object. Test to see if the file is
    already open, if so return an exception. This function is useful to run before any
    serious works starts because if the script tries to open a file that's already open,
    say in Excel, this causes the script to halt with an exception and the work done
    up to that point is lost.

    :param filename_path: Excel filename (Path)
    :return: Excel filename (str)
    """

    excel_output_filename: str = str(
        filename_path.parent / f"{filename_path.stem}_ALL_RESULTS.xlsx"
    )

    try:
        with open(excel_output_filename, "w"):
            pass
    except IOError:
        print(
            f"{Fore.YELLOW}Could not open '{excel_output_filename}', close it "
            + f"if it's already open!{Style.RESET_ALL}"
        )
        sys.exit()

    return excel_output_filename


def write_excel_results_file(
    excel_output_fname: str,
    models: Models,
    mrr: Mrr,
    linear: Linear,
    spiral: Spiral,
    no_spiral: bool = False,
    logger: Callable = print,
):
    """
    Write the analysis results to the output Excel file from a dictionary of
    key:value pairs, where the keys are the Excel file column header text strings
    and the values are the corresponding column data arrays

    :param excel_output_fname:
    :param models:
    :param mrr:
    :param linear:
    :param spiral:
    :param no_spiral:
    :param logger:
    :return: None
    """

    # Create Excel workbook
    wb = Workbook()

    # Save the MMR data to a sheet
    mrr_data_dict = {
        "R (um)": models.R,
        "neff": mrr.neff,
        "max(S) (RIU-1)": mrr.S,
        "Se": mrr.Se,
        "Snr (RIU-1)": mrr.Snr,
        "a2": mrr.a2,
        "tau": mrr.tau,
        "contrast": mrr.contrast,
        f"{models.core_u_name} (um)": mrr.u,
        "gamma (%)": mrr.gamma,
        "Finesse": mrr.Finesse,
        "Q": mrr.Q,
        "FWHM (um)": mrr.FWHM,
        "FSR (um)": mrr.FSR,
    }
    mrr_data: np.ndarray = np.asarray(list(mrr_data_dict.values())).T
    mrr_sheet = wb["Sheet"]
    mrr_sheet.title = "MRR"
    mrr_sheet.append(list(mrr_data_dict.keys()))
    for row in mrr_data:
        mrr_sheet.append(row.tolist())

    # Save the Re(gamma) & Rw(gamma) arrays to a sheet
    ReRw_sheet = wb.create_sheet("Re and Rw")
    ReRw_sheet.append(
        [
            "gamma (%)",
            f"{models.core_u_name} (um)",
            "Re (um)",
            "Rw (um)",
            "A (um-1)",
            "B (um-1)",
        ]
    )
    for line in zip(
        mrr.gamma_resampled * 100, mrr.u_resampled, mrr.Re, mrr.Rw, mrr.A, mrr.B
    ):
        ReRw_sheet.append(line)

    # Save the linear waveguide data to a sheet
    linear_data_dict = {
        "R (um)": models.R,
        "max(S) (RIU-1)": linear.S,
        f"{models.core_u_name} (um)": linear.u,
        "gamma (%)": linear.gamma,
        "L (um)": 2 * models.R,
        "a2": linear.a2,
    }
    linear_data: np.ndarray = np.asarray(list(linear_data_dict.values())).T
    linear_sheet = wb.create_sheet("Linear")
    linear_sheet.append(list(linear_data_dict.keys()))
    for row in linear_data:
        linear_sheet.append(row.tolist())

    # If required, save the spiral data to a sheet
    if not no_spiral:
        spiral_data_dict = {
            "R (um)": models.R,
            "max(S) (RIU-1)": spiral.S,
            f"{models.core_u_name} (um)": spiral.u,
            "gamma (%)": spiral.gamma,
            "n revs (inner+outer)": spiral.n_turns*2,
            "Rmin (um)": spiral.outer_spiral_r_min,
            "L (um)": spiral.L,
            "a2": spiral.a2,
        }
        spiral_data: np.ndarray = np.asarray(list(spiral_data_dict.values())).T
        spiral_sheet = wb.create_sheet("Spiral")
        spiral_sheet.append(list(spiral_data_dict.keys()))
        for row in spiral_data:
            spiral_sheet.append(row.tolist())

    # Save the Excel file to disk
    wb.save(filename=excel_output_fname)
    logger(f"Wrote '{excel_output_fname}'.")


def write_image_data_to_Excel(
    filename: str,
    X: np.ndarray,
    x_label: str,
    Y: np.ndarray,
    y_label: str,
    Zs: list,
    z_labels: list,
):
    """
    Write image data to Excel file
    """

    wb = Workbook()

    # X sheet
    X_sheet = wb["Sheet"]
    X_sheet.title = x_label
    X_sheet.append(X.tolist())

    # Y sheet
    Y_sheet = wb.create_sheet(y_label)
    for y in Y:
        Y_sheet.append([y])

    # Z sheets
    for i, Z in enumerate(Zs):
        Z_sheet = wb.create_sheet(z_labels[i])
        for z in Z:
            Z_sheet.append(z.tolist())

    # Save file
    wb.save(filename=filename)
