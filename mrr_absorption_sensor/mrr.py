"""mrr.py

Micro-ring resonator sensor class

"""
__all__ = ["Mrr"]

from math import e
from pathlib import Path
from typing import Callable

import matplotlib.pyplot as plt
from matplotlib.collections import QuadMesh
from matplotlib.axes import Axes
import numpy as np
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from rich import print
from scipy import optimize
from scipy.special import lambertw

from .constants import CONSTANTS, LINE_STYLES
from .linear import Linear
from .models import Models
from .spiral import Spiral

#
# Type aliasing
#
tuple_of_21_floats = tuple[
    float,
    float,
    float,
    float,
    float,
    float,
    float,
    float,
    float,
    float,
    float,
    float,
    float,
    float,
    float,
    float,
    float,
    float,
    float,
    float,
    float,
]


class Mrr:
    """
    Micro-ring resonator class

    See "Silicon micro-ring resonators" [Bogaerts, 2012] for formulas for Q (20)
    and finesse (21), with Q = (2π * neff * L / lambda) * (F/2π). Q is also the total
    number of field oscillations in the ring x 2π.

    Exposed methods:
        - analyze()
        - plot_optimization_results()
        - plot_combined_linear_mrr_spiral_optimization_results()

    """

    def __init__(self, models: Models, logger: Callable = print):
        """

        Args:
            models (Models): Models class object containing the problem data
            logger (Callable): console logger
        """

        # Load class instance input parameters
        self.models: Models = models
        self.logger: Callable = logger

        # Define class instance internal variables
        self.previous_solution: float = -1

        # Define class instance result variables and arrays
        self.α_bend_a: np.ndarray = np.ndarray([])
        self.α_bend_b: np.ndarray = np.ndarray([])
        self.α_bend: np.ndarray = np.ndarray([])
        self.α_wg: np.ndarray = np.ndarray([])
        self.α_prop: np.ndarray = np.ndarray([])
        self.α: np.ndarray = np.ndarray([])
        self.αl: np.ndarray = np.ndarray([])
        self.wg_a2: np.ndarray = np.ndarray([])
        self.er: np.ndarray = np.ndarray([])
        self.contrast: np.ndarray = np.ndarray([])
        self.finesse: np.ndarray = np.ndarray([])
        self.fsr: np.ndarray = np.ndarray([])
        self.fwhm: np.ndarray = np.ndarray([])
        self.gamma: np.ndarray = np.ndarray([])
        self.gamma_resampled: np.ndarray = np.ndarray([])
        self.max_s: float = 0
        self.max_s_radius: float = 0
        self.n_eff: np.ndarray = np.ndarray([])
        self.plotting_extrema: dict = {}
        self.q: np.ndarray = np.ndarray([])
        self.results: list = []
        self.s: np.ndarray = np.ndarray([])
        self.r_e: np.ndarray = np.ndarray([])
        self.r_w: np.ndarray = np.ndarray([])
        self.s_e: np.ndarray = np.ndarray([])
        self.s_nr: np.ndarray = np.ndarray([])
        self.t_max: np.ndarray = np.ndarray([])
        self.t_min: np.ndarray = np.ndarray([])
        self.tau: np.ndarray = np.ndarray([])
        self.u: np.ndarray = np.ndarray([])
        self.u_resampled: np.ndarray = np.ndarray([])

    #
    # Plotting
    #

    def _calculate_plotting_extrema(self) -> None:
        """
        Calculate the "Se" and "Finesse" plotting maxima so that
        separate plots have the same y axis dynamic range.

        Returns: None

        """

        # Other extrema for Mrr plots
        self.plotting_extrema["Se_plot_max"] = (
            np.ceil(np.amax(self.s_e * np.sqrt(self.wg_a2)) * 1.1 / 10) * 10
        )
        self.plotting_extrema["Finesse_plot_max"] = (
            np.ceil(np.amax(self.finesse / (2 * np.pi)) * 1.1 / 10) * 10
        )

        # Explicit None return
        return None

    @staticmethod
    def _write_image_data_to_excel(
        filename: str,
        x_array: np.ndarray,
        x_label: str,
        y_array: np.ndarray,
        y_label: str,
        z_array: list,
        z_labels: list,
    ) -> None:
        """
        Write xyz image data to Excel file

        Args:
            filename (str): output filename
            x_array (np.ndarray): pixel x coordinates
            x_label (str): x axis label
            y_array (np.ndarray): pixel y coordinates
            y_label (str): y axis label
            z_array (np.ndarray): pixel z coordinates
            z_labels (str): z axis label

        Returns: None

        """

        wb: Workbook = Workbook()

        # X sheet
        x_sheet: Worksheet = wb["Sheet"]
        x_sheet.title = x_label
        x_sheet.append(x_array.tolist())

        # Y sheet
        y_sheet: Worksheet = wb.create_sheet(y_label)
        for y in y_array:
            y_sheet.append([y])

        # Z sheets
        for i, Z in enumerate(z_array):
            z_sheet = wb.create_sheet(z_labels[i])
            for z in Z:
                z_sheet.append(z.tolist())

        # Save file
        wb.save(filename=filename)

        # explicit None return
        return None

    def _plot_mrr_result_2d_maps(self) -> None:
        """
        Plot 2D maps of MRR parameters (S, Snr, Se, u, gamma_fluid, a2, alpha_wg, etc.)
        as a function of (r, u) and (r, gamma), and save to file

        Returns: None

        """

        # Generate 2D map row/column index R,u arrays (x/y)
        r_2d_map_indices: np.ndarray = np.linspace(
            np.log10(self.models.r[0]),
            np.log10(self.models.r[-1]),
            self.models.parms.io.map2D_n_grid_points,
        )
        u_2d_map_indices: np.ndarray = np.linspace(
            self.models.parms.limits.u_min,
            self.models.parms.limits.u_max,
            self.models.parms.io.map2D_n_grid_points,
        )

        # Indices for dashed lines at radii for max(Smrr)
        r_max_s_mrr_index: int = int(
            (np.abs(self.models.r - self.max_s_radius)).argmin()
        )
        r_max_s_mrr_u: float = self.u[r_max_s_mrr_index]

        #
        # 2D maps as a function of r/u
        #

        # 2D map of S(u, R)
        s_2d_map: np.ndarray = np.asarray(
            [
                [
                    self._calc_sensitivity(r=10**log10_R, u=u)
                    for log10_R in r_2d_map_indices
                ]
                for u in u_2d_map_indices
            ]
        )
        fig, ax = plt.subplots()
        cm: QuadMesh = ax.pcolormesh(
            r_2d_map_indices,
            u_2d_map_indices,
            s_2d_map,
            cmap=self.models.parms.io.map2D_colormap,
        )
        ax.invert_yaxis()
        ax.set(
            title="MRR sensitivity as a function of "
            f"{self.models.parms.wg.u_coord_name} and R\n"
            f"{self.models.parms.wg.polarization}"
            f", λ = {self.models.parms.wg.lambda_resonance:.3f} μm"
            f", {self.models.parms.wg.v_coord_name} = "
            f"{self.models.parms.wg.v_coord_value:.3f} μm",
            xlabel="log(R) (μm)",
            ylabel=f"{self.models.parms.wg.u_coord_name} (μm)",
        )
        fig.colorbar(cm, label=r"S (RIU $^{-1}$)")
        max_s_u_min_index: int = np.where(self.s > self.max_s / 25)[0][0]
        ax.plot(
            np.log10(self.models.r[max_s_u_min_index:]),
            self.u[max_s_u_min_index:],
            color=self.models.parms.io.map2D_overlay_color_light,
            label=r"max$\{S(h, R)\}$",
        )
        max_s_u_line_search: np.ndarray = u_2d_map_indices[np.argmax(s_2d_map, axis=0)]
        max_s_u_line_search_min_index: int = np.where(
            np.amax(s_2d_map, axis=0) > self.max_s / 25
        )[0][0]
        ax.plot(
            r_2d_map_indices[max_s_u_line_search_min_index:],
            max_s_u_line_search[max_s_u_line_search_min_index:],
            "k--",
            label=r"max$\{S(h, R)\}$ - line search",
        )
        ax.plot(
            [r_2d_map_indices[0], np.log10(self.max_s_radius)],
            [r_max_s_mrr_u, r_max_s_mrr_u],
            color=self.models.parms.io.map2D_overlay_color_light,
            linestyle=LINE_STYLES["loosely dashdotted"],
            label=rf"max{{max{{$S_{{MRR}}$}}}} = {self.max_s:.0f} RIU $^{{-1}}$"
            f" @ R = {self.max_s_radius:.0f} μm"
            f", {self.models.parms.wg.u_coord_name} = {r_max_s_mrr_u:.3f} μm",
        )
        ax.plot(
            np.log10(self.r_e),
            self.u_resampled,
            color=self.models.parms.io.map2D_overlay_color_light,
            linestyle="--",
            label=r"Re$(\Gamma_{fluid})$",
        )
        ax.plot(
            np.log10(self.r_w),
            self.u_resampled,
            color=self.models.parms.io.map2D_overlay_color_light,
            linestyle="-.",
            label=r"Rw$(\Gamma_{fluid})$",
        )
        ax.legend(loc="lower right")
        name: str = (
            f"{self.models.filename_path.stem}_MRR_2DMAP_S_VS_"
            f"{self.models.parms.wg.u_coord_name}_and_R.png"
        )
        filename: Path = self.models.filename_path.parent / name
        fig.savefig(filename)
        self.logger(f"Wrote '{filename}'.")
        if self.models.parms.io.write_excel_files:
            fname = (
                f"{self.models.filename_path.stem}_MRR_2DMAPS_VS_"
                f"{self.models.parms.wg.u_coord_name}_and_R.xlsx"
            )
            self._write_image_data_to_excel(
                filename=str(self.models.filename_path.parent / fname),
                x_array=10**r_2d_map_indices,
                x_label="R (um)",
                y_array=u_2d_map_indices,
                y_label=f"{self.models.parms.wg.u_coord_name} (um)",
                z_array=[s_2d_map],
                z_labels=["S (RIU-1)"],
            )
            self.logger(f"Wrote '{filename.with_suffix('.xlsx')}'.")

        #
        # 2D maps as a function of R/gamma
        #

        # Generate gamma(u) array matching u array. If the values are not monotonically
        # decreasing due to positive curvature of the modeled values at the beginning of
        # the array, flag as warning and replace values, else pcolormesh() complains.
        gamma_2d_map: np.ndarray = np.asarray(
            [self.models.gamma_of_u(u) * 100 for u in u_2d_map_indices]
        )
        if np.any(np.diff(gamma_2d_map) > 0):
            gamma_2d_map[: int(np.argmax(gamma_2d_map))] = gamma_2d_map[
                int(np.argmax(gamma_2d_map))
            ]
            gamma_2d_map[int(np.argmin(gamma_2d_map)) :] = gamma_2d_map[
                int(np.argmin(gamma_2d_map))
            ]
            self.logger(
                f"[yellow]WARNING! Gamma({self.models.parms.wg.u_coord_name}) is not "
                + "monotonically decreasing, first/last values replaced"
                + " with gamma max/min!"
            )

        # Indices for dashed lines at radii for max(Smrr)
        r_max_s_mrr_gamma: float = self.gamma[r_max_s_mrr_index]

        # 2D map of Smrr(gamma, R)
        fig, ax = plt.subplots()
        cm = ax.pcolormesh(
            r_2d_map_indices,
            gamma_2d_map,
            s_2d_map,
            cmap=self.models.parms.io.map2D_colormap,
        )
        ax.set(
            title=r"MRR sensitivity, $S_{MRR}$, as a function of $\Gamma_{fluid}$ "
            f"and R\n"
            f"{self.models.parms.wg.polarization}, "
            f"λ = {self.models.parms.wg.lambda_resonance:.3f} μm, "
            f"{self.models.parms.wg.v_coord_name} = "
            f"{self.models.parms.wg.v_coord_value:.3f} μm",
            xlabel="log(R) (μm)",
            ylabel=r"$\Gamma_{fluid}$ ($\%$)",
        )
        fig.colorbar(cm, label=r"$S_{MRR}$ (RIU $^{-1}$)")
        ax.plot(
            np.log10(self.models.r),
            self.gamma,
            color=self.models.parms.io.map2D_overlay_color_light,
            label=r"max$\{S_{MRR}(\Gamma_{fluid}, R)\}$",
        )
        ax.plot(
            [r_2d_map_indices[0], np.log10(self.max_s_radius)],
            [r_max_s_mrr_gamma, r_max_s_mrr_gamma],
            color=self.models.parms.io.map2D_overlay_color_light,
            linestyle=LINE_STYLES["loosely dashdotted"],
            label=rf"max{{max{{$S_{{MRR}}$}}}} = {self.max_s:.0f} RIU$^{{-1}}$"
            f" @ R = {self.max_s_radius:.0f} μm"
            rf", $\Gamma$ = {r_max_s_mrr_gamma:.0f}$\%$",
        )
        ax.plot(
            np.log10(self.r_e),
            self.gamma_resampled * 100,
            color=self.models.parms.io.map2D_overlay_color_light,
            linestyle="--",
            label=r"Re$(\Gamma_{fluid})$",
        )
        ax.plot(
            np.log10(self.r_w),
            self.gamma_resampled * 100,
            color=self.models.parms.io.map2D_overlay_color_light,
            linestyle="-.",
            label=r"Rw$(\Gamma_{fluid})$",
        )
        ax.set(
            xlim=(
                np.log10(self.models.plotting_extrema.r_min),
                np.log10(self.models.plotting_extrema.r_max),
            ),
            ylim=(gamma_2d_map[-1], gamma_2d_map[0]),
        )
        ax.legend(loc="lower right")
        filename = (
            self.models.filename_path.parent
            / f"{self.models.filename_path.stem}_MRR_2DMAP_S_VS_GAMMA_and_R.png"
        )
        fig.savefig(filename)
        self.logger(f"Wrote '{filename}'.")

        # 2D map of Snr(gamma, R)
        s_nr_2d_map: np.ndarray = np.asarray(
            [
                [self._calc_s_nr(r=10**log10_R, u=u) for log10_R in r_2d_map_indices]
                for u in u_2d_map_indices
            ]
        )
        fig, ax = plt.subplots()
        cm = ax.pcolormesh(
            r_2d_map_indices,
            gamma_2d_map,
            s_nr_2d_map,
            cmap=self.models.parms.io.map2D_colormap,
        )
        ax.plot(
            np.log10(self.models.r),
            self.gamma,
            color=self.models.parms.io.map2D_overlay_color_dark,
            label=r"max$\{S_{MRR}\}$",
        )
        ax.set(
            title=r"MRR $S_{NR}$ as a function of $\Gamma_{fluid}$ and $R$"
            f"\n{self.models.parms.wg.polarization}"
            f", λ = {self.models.parms.wg.lambda_resonance:.3f} μm"
            f", {self.models.parms.wg.v_coord_name} = "
            f"{self.models.parms.wg.v_coord_value:.3f} μm",
            xlabel="log(R) (μm)",
            ylabel=r"$\Gamma_{fluid}$ ($\%$)",
            xlim=(
                np.log10(self.models.plotting_extrema.r_min),
                np.log10(self.models.plotting_extrema.r_max),
            ),
            ylim=(gamma_2d_map[-1], gamma_2d_map[0]),
        )
        fig.colorbar(cm, label=r"$S_{NR}$ (RIU$^{-1}$)")
        ax.legend(loc="lower right")
        filename = (
            self.models.filename_path.parent
            / f"{self.models.filename_path.stem}_MRR_2DMAP_Snr_VS_GAMMA_and_R.png"
        )
        fig.savefig(filename)
        self.logger(f"Wrote '{filename}'.")

        # 2D map of Se(gamma, R)
        s_e_2d_map: np.ndarray = np.asarray(
            [
                [self._calc_s_e(r=10**log10_R, u=u) for log10_R in r_2d_map_indices]
                for u in u_2d_map_indices
            ]
        )
        fig, ax = plt.subplots()
        cm = ax.pcolormesh(
            r_2d_map_indices,
            gamma_2d_map,
            s_e_2d_map,
            cmap=self.models.parms.io.map2D_colormap,
        )
        ax.plot(
            np.log10(self.models.r),
            self.gamma,
            color=self.models.parms.io.map2D_overlay_color_dark,
            label=r"max$\{S_{MRR}\}$",
        )
        ax.set(
            title=r"MRR $S_e$ as a function of $\Gamma_{fluid}$ and $R$"
            f"\n{self.models.parms.wg.polarization}"
            f", λ = {self.models.parms.wg.lambda_resonance:.3f} μm"
            f", {self.models.parms.wg.v_coord_name} = "
            f"{self.models.parms.wg.v_coord_value:.3f} μm",
            xlabel="log(R) (μm)",
            ylabel=r"$\Gamma_{fluid}$ ($\%$)",
            xlim=(
                np.log10(self.models.plotting_extrema.r_min),
                np.log10(self.models.plotting_extrema.r_max),
            ),
            ylim=(gamma_2d_map[-1], gamma_2d_map[0]),
        )
        fig.colorbar(cm, label=r"$S_e$")
        ax.legend(loc="lower right")
        filename = (
            self.models.filename_path.parent
            / f"{self.models.filename_path.stem}_MRR_2DMAP_Se_VS_GAMMA_and_R.png"
        )
        fig.savefig(filename)
        self.logger(f"Wrote '{filename}'.")

        # 2D map of Se*a(gamma, R)
        s_e_times_a_2d_map: np.ndarray = np.asarray(
            [
                [
                    self._calc_s_e(r=10**log10_R, u=u)
                    * np.sqrt(self._calc_wg_a2(r=10**log10_R, u=u))
                    for log10_R in r_2d_map_indices
                ]
                for u in u_2d_map_indices
            ]
        )
        fig, ax = plt.subplots()
        cm = ax.pcolormesh(
            r_2d_map_indices,
            gamma_2d_map,
            s_e_times_a_2d_map,
            cmap=self.models.parms.io.map2D_colormap,
        )
        ax.plot(
            np.log10(self.models.r),
            self.gamma,
            color=self.models.parms.io.map2D_overlay_color_dark,
            label=r"max$\{S_{MRR}\}$",
        )
        ax.set(
            title=r"MRR $S_e \times a$ as a function of $\Gamma_{fluid}$ and $R$"
            f"\n{self.models.parms.wg.polarization}, "
            f"λ = {self.models.parms.wg.lambda_resonance:.3f} μm, "
            f"{self.models.parms.wg.v_coord_name} = "
            f"{self.models.parms.wg.v_coord_value:.3f} μm",
            xlabel="log(R) (μm)",
            ylabel=r"$\Gamma_{fluid}$ ($\%$)",
            xlim=(
                np.log10(self.models.plotting_extrema.r_min),
                np.log10(self.models.plotting_extrema.r_max),
            ),
            ylim=(gamma_2d_map[-1], gamma_2d_map[0]),
        )
        fig.colorbar(cm, label=r"$S_e \times a$")
        ax.legend(loc="lower right")
        filename = (
            self.models.filename_path.parent
            / f"{self.models.filename_path.stem}_MRR_2DMAP_Se_x_a_VS_GAMMA_and_R.png"
        )
        fig.savefig(filename)
        self.logger(f"Wrote '{filename}'.")

        # 2D map of a2(gamma, R)
        wg_a2_map: np.ndarray = np.asarray(
            [
                [self._calc_wg_a2(r=10**log10_R, u=u) for log10_R in r_2d_map_indices]
                for u in u_2d_map_indices
            ]
        )
        fig, ax = plt.subplots()
        cm = ax.pcolormesh(
            r_2d_map_indices,
            gamma_2d_map,
            wg_a2_map,
            cmap=self.models.parms.io.map2D_colormap,
        )
        ax.plot(
            np.log10(self.models.r),
            self.gamma,
            color=self.models.parms.io.map2D_overlay_color_light,
            label=r"max$\{S_{MRR}\}$",
        )
        ax.plot(
            np.log10(self.r_e),
            self.gamma_resampled * 100,
            color=self.models.parms.io.map2D_overlay_color_light,
            linestyle="--",
            label=r"Re$(\Gamma_{fluid})$",
        )
        ax.plot(
            np.log10(self.r_w),
            self.gamma_resampled * 100,
            color=self.models.parms.io.map2D_overlay_color_light,
            linestyle="-.",
            label=r"Rw$(\Gamma_{fluid})$",
        )
        ax.set(
            title=r"MRR $a^2$ as a function of $\Gamma_{fluid}$ and $R$"
            f"\n{self.models.parms.wg.polarization}, "
            f"λ = {self.models.parms.wg.lambda_resonance:.3f} μm, "
            f"{self.models.parms.wg.v_coord_name} = "
            f"{self.models.parms.wg.v_coord_value:.3f} μm",
            xlabel="log(R) (μm)",
            ylabel=r"$\Gamma_{fluid}$ ($\%$)",
            xlim=(
                np.log10(self.models.plotting_extrema.r_min),
                np.log10(self.models.plotting_extrema.r_max),
            ),
            ylim=(gamma_2d_map[-1], gamma_2d_map[0]),
        )
        fig.colorbar(cm, label=r"$a^2$")
        ax.legend(loc="lower right")
        filename = (
            self.models.filename_path.parent
            / f"{self.models.filename_path.stem}_MRR_2DMAP_a2_VS_GAMMA_and_R.png"
        )
        fig.savefig(filename)
        self.logger(f"Wrote '{filename}'.")

        # 2D map of alpha*L(gamma, R)
        scale_to_db: float = 10 * np.log10(np.e)
        αl_2d_map: np.ndarray = (
            np.asarray(
                [
                    [
                        self._calc_α_l(r=10**log10_R, u=u)
                        for log10_R in r_2d_map_indices
                    ]
                    for u in u_2d_map_indices
                ]
            )
            * scale_to_db
        )
        fig, ax = plt.subplots()
        cm = ax.pcolormesh(
            r_2d_map_indices,
            gamma_2d_map,
            αl_2d_map,
            cmap=self.models.parms.io.map2D_colormap,
        )
        ax.plot(
            np.log10(self.models.r),
            self.gamma,
            color=self.models.parms.io.map2D_overlay_color_dark,
            label=r"max$\{S_{MRR}\}$",
        )
        ax.plot(
            np.log10(self.r_e),
            self.gamma_resampled * 100,
            color=self.models.parms.io.map2D_overlay_color_dark,
            linestyle="--",
            label=r"Re$(\Gamma_{fluid})$",
        )
        ax.plot(
            np.log10(self.r_w),
            self.gamma_resampled * 100,
            color=self.models.parms.io.map2D_overlay_color_dark,
            linestyle="-.",
            label=r"Rw$(\Gamma_{fluid})$",
        )
        ax.set(
            title=r"MRR $\alpha L$ as a function of $\Gamma_{fluid}$ and $R$"
            f"\n{self.models.parms.wg.polarization}, "
            f"λ = {self.models.parms.wg.lambda_resonance:.3f} μm, "
            f"{self.models.parms.wg.v_coord_name} = "
            f"{self.models.parms.wg.v_coord_value:.3f} μm",
            xlabel="log(R) (μm)",
            ylabel=r"$\Gamma_{fluid}$ ($\%$)",
            xlim=(
                np.log10(self.models.plotting_extrema.r_min),
                np.log10(self.models.plotting_extrema.r_max),
            ),
            ylim=(gamma_2d_map[-1], gamma_2d_map[0]),
        )
        fig.colorbar(cm, label=r"$\alpha L$ (dB)")
        ax.legend(loc="lower right")
        filename = (
            self.models.filename_path.parent
            / f"{self.models.filename_path.stem}_MRR_2DMAP_alphaL_VS_GAMMA_and_R.png"
        )
        fig.savefig(filename)
        self.logger(f"Wrote '{filename}'.")

        # 2D map of 1/alpha(gamma, R)
        α_inv_2d_map: np.ndarray = np.asarray(
            [
                [
                    1 / (self._α_prop(u=u) + self.models.α_bend(r=10**log10_R, u=u))
                    for log10_R in r_2d_map_indices
                ]
                for u in u_2d_map_indices
            ]
        )
        fig, ax = plt.subplots()
        cm = ax.pcolormesh(
            r_2d_map_indices,
            gamma_2d_map,
            α_inv_2d_map,
            cmap=self.models.parms.io.map2D_colormap,
        )
        ax.plot(
            np.log10(self.models.r),
            self.gamma,
            color=self.models.parms.io.map2D_overlay_color_dark,
            label=r"max$\{S_{MRR}\}$",
        )
        ax.plot(
            np.log10(self.r_e),
            self.gamma_resampled * 100,
            color=self.models.parms.io.map2D_overlay_color_dark,
            linestyle="--",
            label=r"Re$(\Gamma_{fluid})$",
        )
        ax.plot(
            np.log10(self.r_w),
            self.gamma_resampled * 100,
            color=self.models.parms.io.map2D_overlay_color_dark,
            linestyle="-.",
            label=r"Rw$(\Gamma_{fluid})$",
        )
        ax.set(
            title=r"MRR 1/$\alpha$ as a function of $\Gamma_{fluid}$ and $R$"
            f"\n{self.models.parms.wg.polarization}, "
            f"λ = {self.models.parms.wg.lambda_resonance:.3f} μm, "
            f"{self.models.parms.wg.v_coord_name} = "
            f"{self.models.parms.wg.v_coord_value:.3f} μm",
            xlabel="log(R) (μm)",
            ylabel=r"$\Gamma_{fluid}$ ($\%$)",
            xlim=(
                np.log10(self.models.plotting_extrema.r_min),
                np.log10(self.models.plotting_extrema.r_max),
            ),
            ylim=(gamma_2d_map[-1], gamma_2d_map[0]),
        )
        fig.colorbar(cm, label=r"1/$\alpha$ ($\mu$m)")
        ax.legend(loc="lower right")
        filename = (
            self.models.filename_path.parent
            / f"{self.models.filename_path.stem}_MRR_2DMAP_alpha_inv_VS_GAMMA_and_R.png"
        )
        fig.savefig(filename)
        self.logger(f"Wrote '{filename}'.")

        # Save 2D map data as a function of gamma and R to output Excel file
        if self.models.parms.io.write_excel_files:
            # Construct additional (non-plotted) 2d maps
            α_2d_map: np.ndarray = np.asarray(
                [
                    [
                        self._α_prop(u=u) + self.models.α_bend(r=10**log10_R, u=u)
                        for log10_R in r_2d_map_indices
                    ]
                    for u in u_2d_map_indices
                ]
            )
            α_prop_2d_map: np.ndarray = np.tile(
                [self._α_prop(u=u) for u in u_2d_map_indices],
                (r_2d_map_indices.size, 1),
            ).T
            α_prop_l_2d_map: np.ndarray = (
                np.asarray(
                    [
                        [
                            self._calc_α_prop_l(r=10**log10_R, u=u)
                            for log10_R in r_2d_map_indices
                        ]
                        for u in u_2d_map_indices
                    ]
                )
                * scale_to_db
            )
            α_bend_2d_map: np.ndarray = np.asarray(
                [
                    [
                        self.models.α_bend(r=10**log10_R, u=u)
                        for log10_R in r_2d_map_indices
                    ]
                    for u in u_2d_map_indices
                ]
            )
            α_bend_l_2d_map: np.ndarray = (
                np.asarray(
                    [
                        [
                            self._calc_α_bend_l(r=10**log10_R, u=u)
                            for log10_R in r_2d_map_indices
                        ]
                        for u in u_2d_map_indices
                    ]
                )
                * scale_to_db
            )

            # Write all 2D maps to single Excel file
            self._write_image_data_to_excel(
                filename=str(
                    self.models.filename_path.parent
                    / f"{self.models.filename_path.stem}_MRR_2DMAPS_VS_GAMMA_and_R.xlsx"
                ),
                x_array=10**r_2d_map_indices,
                x_label="R (um)",
                y_array=gamma_2d_map,
                y_label="gamma (%)",
                z_array=[
                    s_2d_map,
                    s_nr_2d_map,
                    s_e_2d_map,
                    α_2d_map,
                    αl_2d_map,
                    α_bend_2d_map,
                    α_bend_l_2d_map,
                    α_prop_2d_map,
                    α_prop_l_2d_map,
                ],
                z_labels=[
                    "S (RIU-1)",
                    "Snr (RIU-1)",
                    "Se",
                    "alpha (um-1)",
                    "alpha x L (dB)",
                    "alpha_bend (um-1)",
                    "alpha_bend x L (dB)",
                    "alpha_prop (um-1)",
                    "alpha_prop x L (dB)",
                ],
            )
            self.logger(f"Wrote '{filename.with_suffix('.xlsx')}'.")

        # Explicit None return
        return None

    def _plot_mrr_sensing_parameters_at_optimum(self) -> None:
        """
        Plots of MRR sensing parameters (max{S}, Snr, Se, u, gamma_fluid, a2, alpha_wg)
        at optimum (max sensitivity) as a function of r, and save to file

        Returns: None

        """

        # max{S}, S_NR, Se, a, u, gamma, Finesse
        fig, axs = plt.subplots(7)
        fig.suptitle(
            "MRR - Sensing parameters\n"
            f"{self.models.parms.wg.polarization}"
            f", λ = {self.models.parms.wg.lambda_resonance:.3f} μm"
            f", {self.models.parms.wg.v_coord_name} = "
            f"{self.models.parms.wg.v_coord_value:.3f} μm"
            f", Lc = {self.models.parms.ring.coupling_length:.1f} μm\n"
            rf"max{{max{{$S$}}}} = {self.max_s:.0f} (RIU$^{{-1}}$)"
            rf" @ $R$ = {self.max_s_radius:.0f} μm"
        )

        # max{S}
        axs_index: int = 0
        axs[axs_index].loglog(self.models.r, self.s)
        axs[axs_index].plot(
            [self.max_s_radius, self.max_s_radius],
            [100, self.models.plotting_extrema.s_max],
            "r--",
        )
        axs[axs_index].set(
            ylabel=r"max$\{S\}$" + "\n" + r"(RIU$^{-1}$)",
            xlim=(
                self.models.plotting_extrema.r_min,
                self.models.plotting_extrema.r_max,
            ),
            ylim=(100, self.models.plotting_extrema.s_max),
            xticklabels=([]),
        )

        # S_NR @ max{S}
        axs_index += 1
        axs[axs_index].loglog(self.models.r, self.s_nr)
        axs[axs_index].plot(
            [self.max_s_radius, self.max_s_radius],
            [10, self.models.plotting_extrema.s_max],
            "r--",
        )
        axs[axs_index].set(
            ylabel=r"S$_{NR}$" + "\n" + r"(RIU $^{-1}$)",
            xlim=(
                self.models.plotting_extrema.r_min,
                self.models.plotting_extrema.r_max,
            ),
            ylim=(10, self.models.plotting_extrema.s_max),
            xticklabels=([]),
        )

        # Se @ max{S}
        axs_index += 1
        axs[axs_index].semilogx(self.models.r, self.s_e * np.sqrt(self.wg_a2))
        axs[axs_index].plot(
            [self.max_s_radius, self.max_s_radius],
            [0, self.plotting_extrema["Se_plot_max"]],
            "r--",
        )
        axs[axs_index].set(
            ylabel=r"S$_e \times a$",
            xlim=(
                self.models.plotting_extrema.r_min,
                self.models.plotting_extrema.r_max,
            ),
            ylim=(0, self.plotting_extrema["Se_plot_max"]),
            xticklabels=([]),
        )

        # u (h or w) @ max{S}
        axs_index += 1
        axs[axs_index].semilogx(self.models.r, self.u)
        axs[axs_index].plot(
            [self.max_s_radius, self.max_s_radius],
            [
                self.models.plotting_extrema.u_min,
                self.models.plotting_extrema.u_max,
            ],
            "r--",
        )
        axs[axs_index].set(
            ylabel=f"{self.models.parms.wg.u_coord_name} (μm)",
            xlim=(
                self.models.plotting_extrema.r_min,
                self.models.plotting_extrema.r_max,
            ),
            ylim=(
                self.models.plotting_extrema.u_min,
                self.models.plotting_extrema.u_max,
            ),
            xticklabels=([]),
        )

        # Gamma_fluid @ max{S}
        axs_index += 1
        axs[axs_index].semilogx(self.models.r, self.gamma)
        axs[axs_index].plot(
            [self.max_s_radius, self.max_s_radius],
            [
                self.models.plotting_extrema.gamma_min,
                self.models.plotting_extrema.gamma_max,
            ],
            "r--",
        )
        axs[axs_index].set(
            ylabel=r"$\Gamma_{fluide}$ ($\%$)",
            xlim=(
                self.models.plotting_extrema.r_min,
                self.models.plotting_extrema.r_max,
            ),
            ylim=(
                self.models.plotting_extrema.gamma_min,
                self.models.plotting_extrema.gamma_max,
            ),
            xticklabels=([]),
        )

        # a2 @ max{S}
        axs_index += 1
        axs[axs_index].semilogx(self.models.r, self.wg_a2)
        axs[axs_index].plot([self.max_s_radius, self.max_s_radius], [0, 1], "r--")
        axs[axs_index].set(
            ylabel=r"$a^2$",
            xlim=(
                self.models.plotting_extrema.r_min,
                self.models.plotting_extrema.r_max,
            ),
            ylim=(0, 1),
            xticklabels=([]),
        )

        # alpha_wg @ max{S}
        axs_index += 1
        axs[axs_index].semilogx(
            self.models.r,
            np.asarray([self.models.α_wg_of_u(u) for u in self.u])
            * CONSTANTS.per_um_to_db_per_cm,
        )
        axs[axs_index].set(
            ylabel=r"α$_{wg}$",
            xlim=(
                self.models.plotting_extrema.r_min,
                self.models.plotting_extrema.r_max,
            ),
            ylim=(
                np.floor(self.models.α_wg_model.min * CONSTANTS.per_um_to_db_per_cm),
                np.ceil(self.models.α_wg_model.max * CONSTANTS.per_um_to_db_per_cm),
            ),
        )

        axs[axs_index].set_xlabel("Ring radius (μm)")
        filename: Path = (
            self.models.filename_path.parent
            / f"{self.models.filename_path.stem}_MRR_sens_parms.png"
        )
        fig.savefig(filename)
        self.logger(f"Wrote '{filename}'.")

        # Explicit None return
        return None

    def _plot_mrr_ring_parameters_at_optimum(self) -> None:
        """
        Plots of MRR ring parameters (max{S}, Q, Finesse, FWHM, FSR, contrast, etc.)
        at optimum (max sensitivity) as a function of r, and save to file

        Returns: None

        """

        # max{S}, Q, Finesse, FWHM, FSR, contrast
        fig, axs = plt.subplots(6)
        fig.suptitle(
            "MRR - Ring parameters"
            f"\n{self.models.parms.wg.polarization}"
            f", λ = {self.models.parms.wg.lambda_resonance:.3f} μm"
            f", {self.models.parms.wg.v_coord_name} = "
            f"{self.models.parms.wg.v_coord_value:.3f} μm\n"
            rf"max{{max{{$S$}}}} = {self.max_s:.0f} (RIU$^{{-1}}$)"
            rf" @ $R$ = {self.max_s_radius:.0f} μm"
        )
        # max{S}
        axs_index = 0
        axs[axs_index].loglog(self.models.r, self.s)
        axs[axs_index].plot(
            [self.max_s_radius, self.max_s_radius],
            [100, self.models.plotting_extrema.s_max],
            "r--",
        )
        axs[axs_index].set(
            ylabel=r"max$\{S\}$",
            xlim=(
                self.models.plotting_extrema.r_min,
                self.models.plotting_extrema.r_max,
            ),
            ylim=(100, self.models.plotting_extrema.s_max),
            xticklabels=([]),
        )

        # Contrast, tau & a @ max{S}
        axs_index += 1
        axs[axs_index].semilogx(self.models.r, self.tau, color="blue", label="τ")
        axs[axs_index].semilogx(
            self.models.r, np.sqrt(self.wg_a2), color="green", label="a"
        )
        axs[axs_index].semilogx(
            self.models.r, self.contrast, color="red", label="contrast"
        )
        axs[axs_index].plot([self.max_s_radius, self.max_s_radius], [0, 1], "r--")
        axs[axs_index].set(
            xlim=(
                self.models.plotting_extrema.r_min,
                self.models.plotting_extrema.r_max,
            ),
            ylim=(0, 1),
            ylabel=r"Contrast, $a$, $\tau$",
            xticklabels=([]),
        )
        axs[axs_index].legend(loc="upper right")

        # ER @ max{S}
        axs_index += 1
        axs[axs_index].semilogx(self.models.r, self.er, label="Q")
        axs[axs_index].plot(
            [self.max_s_radius, self.max_s_radius],
            [0, np.amax(self.er)],
            "r--",
        )
        axs[axs_index].set(
            xlim=(
                self.models.plotting_extrema.r_min,
                self.models.plotting_extrema.r_max,
            ),
            ylabel="Extinction\nratio\n(dB)",
            xticklabels=([]),
        )

        # Q @ max{S}
        axs_index += 1
        axs[axs_index].loglog(self.models.r, self.q, label="Q")
        axs[axs_index].plot(
            [self.max_s_radius, self.max_s_radius],
            [0, np.amax(self.q)],
            "r--",
        )
        axs[axs_index].set(
            xlim=(
                self.models.plotting_extrema.r_min,
                self.models.plotting_extrema.r_max,
            ),
            ylabel="Q",
            xticklabels=([]),
        )

        # (Finesse/2pi) / Se*a @ max{S}
        axs_index += 1
        axs[axs_index].semilogx(
            self.models.r,
            self.finesse / (2 * np.pi) / (self.s_e * np.sqrt(self.wg_a2)),
        )
        axs[axs_index].set(
            xlim=(
                self.models.plotting_extrema.r_min,
                self.models.plotting_extrema.r_max,
            ),
            ylim=(0, 2.5),
            ylabel=r"$\frac{Finesse/2\pi}{S_e\times a}$",
            xticklabels=([]),
        )

        # FWHM, FSR, Finesse/2pi @ max{S}
        axs_index += 1
        axs[axs_index].loglog(self.models.r, self.fwhm * 1e6, "b", label="FWHM")
        axs[axs_index].loglog(self.models.r, self.fsr * 1e6, "g", label="FSR")
        axs[axs_index].set(
            xlim=(
                self.models.plotting_extrema.r_min,
                self.models.plotting_extrema.r_max,
            ),
            xlabel="Ring radius (μm)",
            ylabel="FWHM and FSR\n(pm)",
        )
        ax_right: Axes = axs[axs_index].twinx()
        ax_right.semilogx(
            self.models.r, self.finesse / (2 * np.pi), "k--", label="Finesse/2π"
        )
        ax_right.set_ylabel("Finesse/2π")
        ax_right.grid(visible=False)
        ax_lines: list = (
            axs[axs_index].get_legend_handles_labels()[0]
            + ax_right.get_legend_handles_labels()[0]
        )
        ax_labels: list = (
            axs[axs_index].get_legend_handles_labels()[1]
            + ax_right.get_legend_handles_labels()[1]
        )
        axs[axs_index].legend(ax_lines, ax_labels, loc="upper right")
        axs[axs_index].patch.set_visible(False)
        ax_right.patch.set_visible(True)
        axs[axs_index].set_zorder(ax_right.get_zorder() + 1)

        # Write figure to file
        filename: Path = (
            self.models.filename_path.parent
            / f"{self.models.filename_path.stem}_MRR_ring_parms.png"
        )
        fig.savefig(filename)
        self.logger(f"Wrote '{filename}'.")

        # Explicit None return
        return None

    def plot_mrr_optimization_results(self) -> None:
        """
        Plot all MRR optimization results

        Returns: None

        """
        self._plot_mrr_sensing_parameters_at_optimum()
        self._plot_mrr_ring_parameters_at_optimum()
        if self.models.parms.io.write_2D_maps:
            self._plot_mrr_result_2d_maps()

        # Explicit None return
        return None

    def plot_combined_sensor_optimization_results(
        self, linear: Linear, spiral: Spiral
    ) -> None:
        """
        Plot all 3 sensor optimization results on common plot, and save to file

        Args:
            linear (Linear): Linear sensor object
            spiral (Spiral): Spiral sensor object

        Returns: None

        """

        # Calculate minimum sensitivity required to detect the minimum resolvable
        # change in ni for a given transmission measurement SNR
        s_min: float = (
            10 ** (-self.models.parms.limits.T_SNR / 10)
            / self.models.parms.limits.min_delta_ni
        )

        # Create plot figure
        fig, ax = plt.subplots()
        ax.set_title(
            "Maximum sensitivity for MRR, linear, and spiral sensors\n"
            if self.models.parms.debug.analyze_spiral
            else "Maximum sensitivity for MRR and linear sensors\n"
            f"{self.models.parms.wg.polarization}, "
            f"λ = {self.models.parms.wg.lambda_resonance:.3f} μm"
            f", {self.models.parms.wg.v_coord_name} = "
            f"{self.models.parms.wg.v_coord_value:.3f} μm"
        )

        # MRR
        ax.set_xlabel("Ring radius (μm)")
        ax.set_ylabel(r"Maximum sensitivity (RIU$^{-1}$)")
        ax.loglog(self.models.r, self.s, color="b", label="MRR")

        # Linear waveguide
        ax.loglog(
            self.models.r,
            linear.s,
            color="g",
            label=r"Linear waveguide ($L = 2R$)",
        )
        ax.loglog(
            [
                self.models.plotting_extrema.r_min,
                self.models.plotting_extrema.r_max,
            ],
            [s_min, s_min],
            "r--",
            label="".join(
                [
                    r"min$\{S\}$ to resolve $\Delta n_{i}$",
                    f" = {self.models.parms.limits.min_delta_ni:.0E} "
                    f"@ SNR = {self.models.parms.limits.T_SNR:.0f} dB",
                ]
            ),
        )
        ax.set(
            xlim=(
                self.models.plotting_extrema.r_min,
                self.models.plotting_extrema.r_max,
            ),
            ylim=(100, 1000000),
        )

        # Spiral and MRR/spiral sensitivity ratio, if required
        if self.models.parms.debug.analyze_spiral:
            ax.loglog(
                self.models.r[spiral.s > 1],
                spiral.s[spiral.s > 1],
                color="k",
                label=f"Spiral (spacing = {self.models.parms.spiral.spacing:.0f} μm"
                f", min turns = {self.models.parms.spiral.turns_min:.2f})",
            )
            ax_right: Axes = ax.twinx()
            ax_right.semilogx(
                self.models.r,
                self.s / spiral.s,
                "k--",
                label=r"max$\{S_{MRR}\}$ / max$\{S_{SPIRAL}\}$",
            )
            ax_right.set(ylabel=r"max$\{S_{MRR}\}$ / max$\{S_{SPIRAL}\}$", ylim=(0, 30))
            ax_right.grid(visible=False)
            ax_lines: list = (
                ax.get_legend_handles_labels()[0]
                + ax_right.get_legend_handles_labels()[0]
            )
            ax_labels: list = (
                ax.get_legend_handles_labels()[1]
                + ax_right.get_legend_handles_labels()[1]
            )
            ax.legend(ax_lines, ax_labels, loc="lower right")
            ax.patch.set_visible(False)
            ax_right.patch.set_visible(True)
            ax.set_zorder(ax_right.get_zorder() + 1)
        else:
            ax.legend(loc="lower right")

        # Save figure
        filename: Path = (
            self.models.filename_path.parent
            / f"{self.models.filename_path.stem}_ALL_RESULTS.png"
        )
        fig.savefig(filename)
        self.logger(f"Wrote '{filename}'.")

        # Explicit None return
        return None

    def _calc_α_bend_a_and_b(self, gamma: float) -> tuple[float, float]:
        """
        Fit A & B model parameters for alpha_bend(r) = A*exp(-B*r) @ gamma

        Args:
            gamma (float): gamma value for the model fit

        Returns: model parameters e**(ln(A)) and -B (floats)

        """

        u: float = self.models.u_of_gamma(gamma=gamma)
        r: np.ndarray = np.arange(
            self.models.r_α_bend_min_interp(u),
            self.models.r_α_bend_max_interp(u),
            (self.models.r_α_bend_max_interp(u) - self.models.r_α_bend_min_interp(u))
            / 10,
        )
        α_bend: np.ndarray = np.asarray([self.models.α_bend(r=r, u=u) for r in r])
        minus_b, ln_a = np.linalg.lstsq(
            a=np.vstack([r, np.ones(len(r))]).T, b=np.log(α_bend), rcond=None
        )[0]

        return np.exp(ln_a), -minus_b

    def _objfun_r_w(
        self, r: float, u: float, α_bend_a: float, α_bend_b: float
    ) -> float:
        """
        Calculate the residual for the current solution of Rw modeled
        using equation (15) in the paper

        Args:
            r (float): radius
            u (float): waveguide core free parameter
            α_bend_a (float): alpha_bend(r) model parameter A
            α_bend_b (float): alpha_bend(r) model parameter B

        Returns: residual (float)

        """

        α_bend: float = α_bend_a * np.exp(-α_bend_b * r)
        residual: float = 1 - r * (2 * np.pi) * (
            self._α_prop(u=u) + (1 - α_bend_b * r) * α_bend
        )

        return residual**2

    def _calc_r_e_and_r_w(self, gamma: float) -> tuple[float, float, float, float]:
        """
        Calculate Re(gamma) and Rw(gamma)

        Args:
            gamma (float): gamma

        Returns: Re(float), Rw(float)
                 A, B (floats, alpha_bend(r) model parameters A & B)

        """

        # u corresponding to gamma
        u: float = self.models.u_of_gamma(gamma=gamma)

        # alpha_bend(R) = A*exp(-BR) model parameters @gamma
        α_bend_a, α_bend_b = self._calc_α_bend_a_and_b(gamma=gamma)

        # Re
        w: float = lambertw(-e * self._α_prop(u=u) / α_bend_a, k=-1).real
        r_e: float = (1 / α_bend_b) * (1 - w)

        # Rw
        optimization_result: optimize.minimize = optimize.minimize(
            fun=self._objfun_r_w,
            x0=np.asarray(r_e),
            args=(u, α_bend_a, α_bend_b),
            method="SLSQP",
        )
        r_w: float = optimization_result["x"][0]

        return r_e, r_w, α_bend_a, α_bend_b

    #
    # Optimization
    #

    def _α_prop(self, u: float) -> float:
        """
        alpha_prop(u) = alpha_wg(u) + gamma_fluid(u)*alpha_fluid

        Args:
            u (float): waveguide core free parameter

        Returns: alpha_prop (float)

        """

        return self.models.α_wg_of_u(u=u) + (
            self.models.gamma_of_u(u) * self.models.α_fluid
        )

    def _calc_α_prop_l(self, r: float, u: float) -> float:
        """
        Propagation loss component of total round-trip losses : alpha_prop(u)*length(r)

        Args:
            r (float): radius
            u (float): waveguide core free parameter

        Returns: alpha_prop(u)*length(r) (float)

        """

        return self._α_prop(u=u) * (
            (2 * np.pi * r) + (2 * self.models.parms.ring.coupling_length)
        )

    def _calc_α_bend_l(self, r: float, u: float) -> float:
        """
        Bending loss component of total round-trip losses: alpha_bend(r, u)*length(r)

        Args:
            r (float): radius
            u (float): waveguide core free parameter

        Returns: alpha_bend(r, u)*length(r) (float)

        """
        return self.models.α_bend(r=r, u=u) * (2 * np.pi * r)

    def _calc_α_l(self, r: float, u: float) -> float:
        """
        Total ring round-trip loss factor: alpha*L = (alpha*_prop + alpha__bend)*length

        Args:
            r (float): radius
            u (float): waveguide core free parameter

        Returns: alpha*L (float)

        """

        return self._calc_α_prop_l(r=r, u=u) + self._calc_α_bend_l(r=r, u=u)

    def _calc_wg_a2(self, r: float, u: float) -> float:
        """
        Ring round trip losses: a2 = e**(-alpha*length)

        Args:
            r (float): radius
            u (float): waveguide core free parameter

        Returns: a2 (float)

        """

        return np.e ** -self._calc_α_l(r=r, u=u)

    def _calc_s_nr(self, r: float, u: float) -> float:
        """
        Calculate "non-resonant" portion of mrr sensitivity Snr (see eq.2 in paper)

        Args:
            r (float): radius
            u (float): waveguide core free parameter

        Returns: Snr (float)

        """
        return (
            (4 * np.pi / self.models.parms.wg.lambda_resonance)
            * ((2 * np.pi * r) + (2 * self.models.parms.ring.coupling_length))
            * self.models.gamma_of_u(u)
            * self._calc_wg_a2(r=r, u=u)
        )

    def _calc_s_e(self, r: float, u: float) -> float:
        """
        Calculate "resonant" portion of mrr sensitivity Se (see eq.2 in paper)

        Args:
            r (float): radius
            u (float): waveguide core free parameter

        Returns: Se (float)

        """

        return (
            2
            / (3 * np.sqrt(3))
            / (np.sqrt(self._calc_wg_a2(r=r, u=u)) * (1 - self._calc_wg_a2(r=r, u=u)))
        )

    def _calc_sensitivity(self, r: float, u: float) -> float:
        """
        Calculate rig sensitivity at radius r for a given core dimension u

        Args:
            r (float): radius
            u (float): waveguide core free parameter

        Returns: sensitivity (float)

        """

        s: float = self._calc_s_nr(r=r, u=u) * self._calc_s_e(r=r, u=u)
        assert s >= 0, "S should not be negative!"

        return s

    def _obj_fun(self, u: float, *args) -> float:
        """
        Objective function for the non-linear minimization in find_max_sensitivity()

        Args:
            u (float): waveguide core free parameter in the optimization
            *args (float): r, radius

        Returns: negative of sensitivity (to maximize sensitivity in the optimization,
                 scaled by 1000 for a reasonable range to aid convergence)

        """

        # Fetch additional parameters
        r: float = args[0]

        # Minimizer sometimes tries values of the solution vector outside the bounds...
        u = min(u, self.models.parms.limits.u_max)
        u = max(u, self.models.parms.limits.u_min)

        # Calculate sensitivity at current solution vector S(r, h)
        s: float = self._calc_sensitivity(r=r, u=u)

        return -s / 1000

    def _find_max_sensitivity(self, r: float) -> tuple_of_21_floats:
        """
        Calculate maximum mrr sensitivity at radius r over all u

        Args:
            r (float): radius

        Returns:
            s (float): ring maximum sensitivity
            u_max_s (float): u @ max{S}
            gamma (float): waveguide gamma @ max{S}
            s_nr (float): ring Snr @ max{S}
            s_e (float): ring Se @ max{S}
            α_bend (float): waveguide alpha_bend @ max{S}
            α_wg (float): waveguide alpha_wg @ max{S}
            α_prop (float): waveguide alpha_prop @ max{S}
            α (float): waveguide a @ max{S}
            αl (float): alpha*length @ max{S}
            wg_a2 (float): ring a2 @ max{S}
            tau (float): ring tau @ max{S}
            t_max (float): ring t-a @ max{S}
            t_min (float): ring t+a @ max{S}
            er (float): ring extinction ratio @ max{S}
            contrast (float): ring contrast @ max{S}
            n_eff (float): neff @ max{S}
            q (float): ring Q @ max{S}
            finesse (float): ring Finesse @ max{S}
            fwhm (float): ring FWHM @ max{S}
            fsr (float): ring FSR @ max{S}

        """

        # Determine u search domain extrema
        u_min, u_max = self.models.u_search_domain(r)

        # If this is the first optimization, set the initial guess for u at the
        # maximum value in the domain (at small radii, bending losses are high,
        # the optimal solution will be at high u), else use previous solution.
        u0: float = u_max if self.previous_solution == -1 else self.previous_solution

        # Find u that maximizes S at radius r.
        if u_min != u_max:
            if self.models.parms.fit.optimization_local:
                optimization_result: optimize.OptimizeResult = optimize.minimize(
                    fun=self._obj_fun,
                    x0=np.asarray([u0]),
                    bounds=((u_min, u_max),),
                    args=(r,),
                    method=self.models.parms.fit.optimization_method,
                    tol=1e-9,
                )
            else:
                optimization_result = optimize.shgo(
                    func=self._obj_fun,
                    bounds=[(u_min, u_max)],
                    args=(r,),
                    iters=3,
                    options={"minimize_every_iter": True},
                )
            u_max_s: float = optimization_result.x[0]
        else:
            u_max_s = u0

        # Update previous solution
        self.previous_solution = u_max_s

        # Calculate sensitivity and other parameters at the solution
        s: float = self._calc_sensitivity(r=r, u=u_max_s)

        # Calculate other useful MRR parameters at the solution
        wg_a2: float = self._calc_wg_a2(r=r, u=u_max_s)
        a: float = np.sqrt(wg_a2)
        tau: float = (np.sqrt(3) * wg_a2 - np.sqrt(3) - 2 * a) / (wg_a2 - 3)
        t_max: float = ((tau + a) / (1 + tau * a)) ** 2
        t_min: float = ((tau - a) / (1 - tau * a)) ** 2
        gamma: float = self.models.gamma_of_u(u_max_s) * 100
        n_eff: float = self.models.n_eff_of_u(u_max_s)
        finesse: float = np.pi * (np.sqrt(tau * a)) / (1 - tau * a)
        q: float = (
            n_eff * (2 * np.pi * r) / self.models.parms.wg.lambda_resonance
        ) * finesse
        fwhm: float = self.models.parms.wg.lambda_resonance / q
        fsr: float = finesse * fwhm
        contrast: float = t_max - t_min
        er: float = 10 * np.log10(t_max / t_min)
        s_nr: float = self._calc_s_nr(r=r, u=u_max_s)
        s_e: float = self._calc_s_e(r=r, u=u_max_s)
        α_bend: float = self.models.α_bend(r=r, u=u_max_s)
        α_wg: float = self.models.α_wg_of_u(u=u_max_s)
        α_prop: float = self._α_prop(u=u_max_s)
        α: float = α_prop + α_bend
        αl: float = self._calc_α_l(r=r, u=u_max_s)

        # Return results to calling program
        return (
            s,
            u_max_s,
            gamma,
            s_nr,
            s_e,
            α_bend,
            α_wg,
            α_prop,
            α,
            αl,
            wg_a2,
            tau,
            t_max,
            t_min,
            er,
            contrast,
            n_eff,
            q,
            finesse,
            fwhm,
            fsr,
        )

    def analyze(self) -> None:
        """
        Analyse the MRR sensor performance for all radii in the r domain,
        store the results in the Mrr object instance

        Returns: None

        """

        # Analyse the sensor performance for all radii in the R domain
        self.results = [self._find_max_sensitivity(r=r) for r in self.models.r]

        # Unpack the analysis results as a function of radius into separate lists, the
        # order must be the same as in the find_max_sensitivity() return statement above
        [
            self.s,
            self.u,
            self.gamma,
            self.s_nr,
            self.s_e,
            self.α_bend,
            self.α_wg,
            self.α_prop,
            self.α,
            self.αl,
            self.wg_a2,
            self.tau,
            self.t_max,
            self.t_min,
            self.er,
            self.contrast,
            self.n_eff,
            self.q,
            self.finesse,
            self.fwhm,
            self.fsr,
        ] = list(np.asarray(self.results).T)

        # Find maximum sensitivity overall and corresponding radius
        self.max_s = np.amax(self.s)
        self.max_s_radius = self.models.r[np.argmax(self.s)]

        # Calculate Re(gamma) and Rw(gamma)
        self.gamma_resampled = np.linspace(
            self.models.parms.limits.gamma_min, self.models.parms.limits.gamma_max, 500
        )
        self.u_resampled = np.asarray(
            [self.models.u_of_gamma(g) for g in self.gamma_resampled]
        )
        self.r_e, self.r_w, self.α_bend_a, self.α_bend_b = np.asarray(
            list(
                zip(
                    *[
                        self._calc_r_e_and_r_w(gamma=gamma)
                        for gamma in self.gamma_resampled
                    ]
                )
            )
        )

        # Calculate extrema for plotting of results
        self._calculate_plotting_extrema()

        # Console message
        self.logger("MRR sensor analysis done.")

        # Explicit None return
        return None
