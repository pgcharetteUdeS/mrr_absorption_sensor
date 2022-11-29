"""fileio.py

File I/O utilities

All lengths are in units of um

"""
__all__ = [
    "load_toml_file",
    "validate_excel_results_file",
    "write_excel_results_file",
]

import dacite
from dacite import from_dict
from dataclasses import asdict
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
from typing import Callable

import numpy as np
import toml
from rich import print

from .constants import CONSTANTS, InputParameters
from .version import __version__


from .linear import Linear
from .models import Models
from .mrr import Mrr
from .spiral import Spiral


def load_toml_file(filename: Path, logger: Callable = print) -> InputParameters:
    """
    Load problem data from .toml file dict into a InputParameters dataclass

    Args:
        filename (Path): .toml input file containing the problem data
        logger (Callable, optional): console logger (optional)

    Returns: parms (InputParameters class object)

    """

    # Load problem parameters from a .toml file into an InputParameters dataclass
    toml_dict: dict = toml.load(filename)
    toml_dict |= {"filename": filename}
    parms: InputParameters = from_dict(
        data_class=InputParameters, data=toml_dict, config=dacite.Config(strict=True)
    )
    logger(f"Loaded information from '{filename}' ({len(parms.geom)} geom entries).")

    # Write out the parameters to a .toml file, for reference purposes
    filename_txt: Path = (
        filename.parent / parms.io.output_sub_dir / f"{filename.stem}_parameters.toml"
    )
    with open(filename_txt, "w") as f:
        f.write(f"# mrr_absorption_sensor package {__version__}\n")
        toml.dump(asdict(parms), f)
    logger(f"Wrote input parameters to '{filename_txt}'.")

    # Return dataclass with problem parameters
    return parms


def validate_excel_results_file(filename_path: Path) -> Path:
    """
    Define output Excel filename string from a Path object. Test to see if the file is
    already open, if so return an exception.

    This function is useful to run before any serious works starts because if the script
    tries to open a file that's already open, say in Excel, this causes the script to
    halt with an exception and the work done up to that point is lost.

    Args:
        filename_path (Path): filename path object to test for openness, conversion

    Returns:
        Path: filename path

    """

    excel_output_filename: Path = filename_path.with_name(
        f"{filename_path.stem}_ALL_RESULTS.xlsx"
    )

    try:
        with open(str(excel_output_filename), "a"):
            pass
    except IOError:
        raise IOError(
            f"Could not open '{excel_output_filename}', close it "
            "if it's already open!"
        ) from None

    return excel_output_filename


def _add_sheet_to_workbook(sensor_dict: dict, wb: Workbook, sensor_name: str):
    """
    Add a sheet to the Workbook with the sensor data
    Args:
        sensor_dict (dict): dictionary containing the sensor data
        wb (Workbook): Workbook to which the sheet is appended
        sensor_name (): Sensor type for naming the sheet

    Returns: None

    """
    sensor_data: np.ndarray = np.asarray(list(sensor_dict.values())).T
    sheet: Worksheet = wb.create_sheet(sensor_name)
    sheet.append(list(sensor_dict.keys()))
    for row in sensor_data:
        sheet.append(row.tolist())
    return None


def write_excel_results_file(
    excel_output_path: Path | None,
    models: Models,
    mrr: Mrr,
    linear: Linear,
    spiral: Spiral,
    parms: InputParameters,
    logger: Callable = print,
) -> None:
    """
    Write the analysis results to the output Excel file from a dictionary of
    key:value pairs, where the keys are the Excel file column header text strings
    and the values are the corresponding column data arrays

    Args:
        excel_output_path (Path): Excel output file Path
        models (Models): Models class object
        mrr (Mrr): Mrr class object
        linear (Linear): Linear class object
        spiral (Spiral): Spiral class object
        parms (InputParameters): InputParameters clas object
        logger (Callable, optional): console logger

    Returns: None

    """
    # Create Excel workbook
    wb: Workbook = Workbook()
    wb.remove_sheet(wb.worksheets[0])

    # Save the MMR data to a sheet
    mrr_dict: dict = {
        "R_um": models.r,
        "neff": mrr.n_eff,
        "maxS_RIU_inv": mrr.s,
        "Se": mrr.s_e,
        "Snr_RIU_inv": mrr.s_nr,
        "alpha_bend_dB_per_cm": mrr.α_bend * CONSTANTS.per_um_to_db_per_cm,
        "alpha_wg_dB_per_cm": mrr.α_wg * CONSTANTS.per_um_to_db_per_cm,
        "alpha_prop_dB_per_cm": mrr.α_prop * CONSTANTS.per_um_to_db_per_cm,
        "alpha_dB_per_cm": mrr.α * CONSTANTS.per_um_to_db_per_cm,
        "alphaL": mrr.αl,
        "a2": mrr.wg_a2,
        "tau": mrr.tau,
        "T_max": mrr.t_max,
        "T_min": mrr.t_min,
        "ER_dB": mrr.er,
        "contrast": mrr.contrast,
        f"{models.parms.wg.u_coord_name}_um": mrr.u,
        "gamma_percent": mrr.gamma,
        "Finesse": mrr.finesse,
        "Q": mrr.q,
        "FWHM_um": mrr.fwhm,
        "FSR_um": mrr.fsr,
    }
    _add_sheet_to_workbook(sensor_dict=mrr_dict, wb=wb, sensor_name="MRR")

    # Save the calculated and interpolated alpha_wg(u) values to a sheet
    alpha_wg_sheet: Worksheet = wb.create_sheet("alpha_wg")
    alpha_wg_interp_sheet: Worksheet = wb.create_sheet("alpha_wg_interp")
    if models.parms.wg.v_coord_name == "w":
        alpha_wg_sheet.append(["height_um", "alpha_wg_dB_per_cm"])
        alpha_wg_interp_sheet.append(["height_um", "alpha_wg_dB_per_cm"])
    else:
        alpha_wg_sheet.append(["width_um", "alpha_wg_dB_per_cm"])
        alpha_wg_interp_sheet.append(["width_um", "alpha_wg_dB_per_cm"])
    for value in models.parms.geom.values():
        alpha_wg_sheet.append([value.u, value.alpha_wg])
    u_data: np.ndarray = np.asarray([value.u for value in models.parms.geom.values()])
    u_interp: np.ndarray = np.linspace(u_data[0], u_data[-1], 100)
    alpha_wg_modeled: np.ndarray = (
        np.asarray([models.α_wg_of_u(u) for u in u_interp])
        * CONSTANTS.per_um_to_db_per_cm
    )
    for u, alpha_wg in zip(u_interp, alpha_wg_modeled):
        alpha_wg_interp_sheet.append([u, alpha_wg])

    # Save the Re(gamma) & Rw(gamma) arrays to a sheet
    re_rw_sheet: Worksheet = wb.create_sheet("Re and Rw")
    re_rw_sheet.append(
        [
            "gamma_percent",
            f"{models.parms.wg.u_coord_name}_um",
            "Re_um",
            "Rw_um",
            "A_um_inv",
            "B_um_inv",
        ]
    )
    for line in zip(
        mrr.gamma_resampled * 100,
        mrr.u_resampled,
        mrr.r_e,
        mrr.r_w,
        mrr.α_bend_a,
        mrr.α_bend_b,
    ):
        re_rw_sheet.append(line)

    # Save the linear waveguide data to a sheet
    linear_dict: dict = {
        "R_um": models.r,
        "maxS_RIU_inv": linear.s,
        f"{models.parms.wg.u_coord_name}_um": linear.u,
        "gamma_percent": linear.gamma,
        "L_um": 2 * models.r,
        "a2": linear.wg_a2,
    }
    _add_sheet_to_workbook(sensor_dict=linear_dict, wb=wb, sensor_name="Linear")

    # If required, save the spiral data to a sheet
    if parms.debug.analyze_spiral:
        spiral_dict: dict = {
            "R_um": models.r,
            "maxS_RIU_inv": spiral.s,
            f"{models.parms.wg.u_coord_name}_um": spiral.u,
            "gamma_percent": spiral.gamma,
            "n_revs": spiral.n_turns * 2,
            "Rmin_um": spiral.outer_spiral_r_min,
            "L_um": spiral.l,
            "a2": spiral.wg_a2,
        }
        _add_sheet_to_workbook(sensor_dict=spiral_dict, wb=wb, sensor_name="Spiral")

    # Save the Excel file to disk
    wb.save(filename=str(excel_output_path))
    logger(f"Wrote '{excel_output_path}'.")

    # Explicit None return
    return None
